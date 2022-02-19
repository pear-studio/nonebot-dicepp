import os
from typing import Dict, List

import openpyxl
from openpyxl.comments import Comment

from utils.logger import dice_log
from utils.localdata import read_xlsx

from core.config.config_item import ConfigItem
from core.config.basic import DATA_PATH
from core.config.common import DEFAULT_CONFIG, DEFAULT_CONFIG_COMMENT

CONFIG_FILE_PATH = "config.xlsx"


class ConfigManager:
    def __init__(self, data_path: str, identifier: str):
        self.data_path = os.path.join(data_path, CONFIG_FILE_PATH)
        self.identifier = identifier
        self.all_configs: Dict[str, ConfigItem] = {}

        # 默认配置
        for key in DEFAULT_CONFIG.keys():
            self.register_config(key, DEFAULT_CONFIG[key], DEFAULT_CONFIG_COMMENT[key])

    def load_config(self):
        """用文档里的配置覆写之前的配置"""
        if not os.path.exists(self.data_path):
            dice_log(f"[BotConfig] [Load] 无法读取配置文件 {self.data_path.replace(DATA_PATH, '~')}")
            return
        workbook = read_xlsx(self.data_path)
        if self.identifier in workbook.sheetnames:
            sheet = workbook[self.identifier]
        elif "Default" in workbook.sheetnames:
            sheet = workbook["Default"]
        else:
            dice_log(f"[BotConfig] [Load] 无法读取配置文件 {self.data_path.replace(DATA_PATH, '~')} {self.identifier}")
            return
        for row in sheet.iter_rows():
            key = str(row[0].value)  # 第一个元素为关键字
            if key not in self.all_configs:
                continue
            comment: str = self.all_configs[key].comment  # 沿用原来的注释, 不用文件里的
            self.all_configs[key] = ConfigItem(key, comment=comment)
            for text in [str(cell.value) for cell in row[1:] if (cell.value is not None)]:
                self.all_configs[key].add(text)
        workbook.close()
        dice_log(f"[BotConfig] [Load] 成功读取配置文件 {self.data_path.replace(DATA_PATH, '~')}")

    def save_config(self):
        """按现在的设置多个机器人会读写同一个配置文件, 如果并行可能存在写冲突, 现在应该是单线程异步, 应该没问题"""
        def save_loc_text_to_row(sheet, item: ConfigItem, row: int):
            header = sheet.cell(row=row, column=1, value=item.key)
            if item.comment:
                header.comment = Comment(item.comment, "DicePP")
            if not item.contents:
                sheet.cell(row=row, column=2, value="")
            else:
                for ci, text in enumerate(item.contents):
                    sheet.cell(row=row, column=ci + 2, value=text)

        if os.path.exists(self.data_path):
            workbook = read_xlsx(self.data_path)
        else:
            workbook = openpyxl.Workbook()
            for name in workbook.sheetnames:
                del workbook[name]
        if self.identifier in workbook.sheetnames:
            cur_sheet = workbook[self.identifier]
        else:
            cur_sheet = workbook.create_sheet(self.identifier)
        for ri, cfg_item in enumerate(self.all_configs.values()):
            save_loc_text_to_row(cur_sheet, cfg_item, ri + 1)

        try:
            workbook.save(self.data_path)
        except PermissionError:
            dice_log(f"[BotConfig] [Save] 无法保存本地化文件 {self.data_path.replace(DATA_PATH, '~')}, 没有写入权限")

        workbook.close()

        dice_log(f"[BotConfig] [Save] 成功更新配置文件 {self.data_path.replace(DATA_PATH, '~')}")

    def register_config(self, key: str, origin_str: str, comment: str = ""):
        """
        将一个配置注册至Helper中
        Args:
            key: 配置关键字
            origin_str: 配置的默认值
            comment: 注释
        """
        self.all_configs[key] = ConfigItem(key, origin_str, comment)

    def get_config(self, key: str) -> List[str]:
        """
        获取本地化语句
        Args:
            key: 本地化语句的关键字
        """
        return self.all_configs[key].get()
