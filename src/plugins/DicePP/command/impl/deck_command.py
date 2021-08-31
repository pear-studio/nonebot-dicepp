"""
命令模板, 复制到新创建的文件里修改
"""

from typing import List, Tuple, Any, Iterable, Set, Dict
import random
import re
import os
import openpyxl
from openpyxl.comments import Comment

import bot_config
from bot_core import Bot
from command.command_config import *
from command.dicepp_command import UserCommandBase, custom_user_command, MessageMetaData, preprocess_msg
from command.bot_command import BotCommandBase, PrivateMessagePort, GroupMessagePort, BotSendMsgCommand
from bot_utils.localdata import read_xlsx, update_xlsx, create_parent_dir
from bot_utils.string import match_substring
from localization import LocalizationHelper
from roll_dice import preprocess_roll_exp, is_roll_exp, exec_roll_exp
from logger import dice_log

LOC_DRAW_RESULT = "draw_result"
LOC_DRAW_RESULT_INLINE = "draw_result_inline"
LOC_DRAW_SINGLE = "draw_single"
LOC_DRAW_MULTI = "draw_multi"
LOC_DRAW_FIN_ALL = "draw_finalize_all"
LOC_DRAW_FIN_INNER = "draw_finalize_inner"
LOC_DRAW_ERR_EMPTY_DECK = "draw_error_empty_deck"
LOC_DRAW_ERR_TIME = "draw_error_time"
LOC_DRAW_ERR_NO_DECK = "draw_error_no_deck"
LOC_DRAW_ERR_VAGUE_DECK = "draw_error_vague_deck"

CFG_DECK_DATA_PATH = "deck_data_path"

DRAW_LIMIT = 10  # 指令抽卡的上限
HLDL_DRAW_LIMIT = 50  # 高级抽卡语言中抽卡的上限

DECK_ITEM_FIELD_CONTENT = "Content"
DECK_ITEM_FIELD_WEIGHT = "Weight"
DECK_ITEM_FIELD_REDRAW = "Redraw"
DECK_ITEM_FIELD_FINAL = "Finalize"

DECK_ITEM_FIELD = [DECK_ITEM_FIELD_CONTENT, DECK_ITEM_FIELD_WEIGHT, DECK_ITEM_FIELD_REDRAW, DECK_ITEM_FIELD_FINAL]
DECK_ITEM_FIELD_COMMENT = {DECK_ITEM_FIELD_CONTENT: "词条内容",
                           DECK_ITEM_FIELD_WEIGHT: "词条权重, 大于等于1的整数, 不填默认为1",
                           DECK_ITEM_FIELD_REDRAW: "抽出后是否放回, 1为放回, 0为不放回, 默认为1",
                           DECK_ITEM_FIELD_FINAL: "抽出后是否结束抽卡, 0为不结束, 1为终止当前层抽取但不终止外层抽取, 2为终止所有抽取. 默认为0."
                           }


class ForceFinal(Exception):
    """强制终止所有抽取"""

    def __init__(self, info: str):
        self.info = info

    def __str__(self):
        return self.info


class DeckItem:
    """牌库中的一个元素"""

    def __init__(self, content: str, weight: int = 1, redraw: bool = True, final_type: int = False):
        """
        Args:
            content: 内容, 包含高级抽卡语言
            weight: 权重
            redraw: 抽出后是否放回牌库中
            final_type: 抽出后是否终止抽牌, 0为不终止, 1为终止当前抽取但不终止外层抽取, 2为终止所有抽取
        """
        self.content = content
        self.weight = weight
        self.redraw = redraw
        self.final_type = final_type
        if self.weight <= 0:
            self.weight = 1

    def get_result(self, decks: Iterable["Deck"], loc_helper: LocalizationHelper, ignore: bool = True) -> str:
        """处理高级抽卡语言"""

        def handle_roll(match):
            roll_exp = preprocess_roll_exp(match.group(1))
            if is_roll_exp(roll_exp):
                roll_res = exec_roll_exp(roll_exp)
                return roll_res.get_complete_result()
            else:
                if ignore:
                    return match.group(1)
                else:
                    raise ValueError(f"{roll_exp} in {match.group()} is an invalid roll expression!")

        def handle_draw(match):
            target_deck_str = match.group(1)
            draw_exp = preprocess_roll_exp(match.group(2)).strip()
            draw_times: int
            draw_times_str: str
            # 得到抽取次数
            try:
                draw_times = int(draw_exp)
                draw_times_str = draw_exp
            except ValueError:
                if is_roll_exp(draw_exp):
                    roll_res = exec_roll_exp(draw_exp)
                    draw_times = roll_res.get_val()
                    draw_times_str = roll_res.get_complete_result()
                else:
                    if ignore:
                        return f"{target_deck_str}*{draw_exp}"
                    else:
                        raise ValueError(f"{draw_exp} in {match.group()} is an invalid roll expression!")
            if draw_times <= 0 or draw_times > HLDL_DRAW_LIMIT:
                if ignore:
                    return f"{target_deck_str}*{draw_times_str}"
                else:
                    raise ValueError(f"{draw_exp} in {match.group()} results an invalid value! value:{draw_times}")
            # 搜索目标牌库
            target_deck = None
            for deck in decks:
                if deck.name == target_deck_str:
                    target_deck = deck
                    break
            if not target_deck:
                if ignore:
                    return f"{target_deck_str}*{draw_times_str}"
                else:
                    raise ValueError(f"{target_deck_str} in {match.group()} is an invalid deck!")
            draw_result = target_deck.draw(draw_times, decks, loc_helper, ignore).replace("\n", " ")  # 嵌套抽取不需要换行
            return loc_helper.format_loc_text(LOC_DRAW_RESULT_INLINE, times=draw_times, deck_name=target_deck_str,
                                              result=draw_result)

        result = self.content.strip()
        result = re.sub(r"ROLL\((.{1,30}?)\)", handle_roll, result)
        result = re.sub(r"DRAW\((.{1,30}?),\s*(.{1,30}?)\)", handle_draw, result)
        if self.final_type == 2:
            raise ForceFinal(result + "\n" + loc_helper.format_loc_text(LOC_DRAW_FIN_ALL))
        return result


class Deck:
    """牌库"""

    def __init__(self, name: str):
        self.name = name
        self.items: List[DeckItem] = []
        self.weight_sum: int = 0

    def add_item(self, item: DeckItem):
        self.items.append(item)
        self.weight_sum += item.weight

    def draw(self, times: int, decks: Iterable["Deck"], loc_helper: LocalizationHelper, ignore: bool = True) -> str:
        weight_sum_cur = self.weight_sum
        index_mask: Set[int] = set()
        feedback: str = ""
        for t in range(times):
            if weight_sum_cur <= 0:  # 牌库被抽光了, 全都是不放回的
                feedback += loc_helper.format_loc_text(LOC_DRAW_ERR_EMPTY_DECK)
                break
            weight_random = random.randint(1, weight_sum_cur)
            item_selected = None
            for i, item in enumerate(self.items):
                if i in index_mask:
                    continue
                weight_random -= item.weight
                if weight_random <= 0:
                    item_selected = item
                    break

            if not item_selected.redraw:  # 抽到的不放回
                index_mask.add(self.items.index(item_selected))
                weight_sum_cur -= item_selected.weight

            try:
                content = item_selected.get_result(decks, loc_helper, ignore)
            except ForceFinal as e:
                if times > 1:
                    feedback += loc_helper.format_loc_text(LOC_DRAW_MULTI, time=t+1, content=e.info + "\n")
                else:
                    feedback += loc_helper.format_loc_text(LOC_DRAW_SINGLE, content=e.info)
                raise ForceFinal(feedback)  # 继续抛出当前结果

            if times > 1:
                feedback += loc_helper.format_loc_text(LOC_DRAW_MULTI, time=t+1, content=content + "\n")
            else:
                feedback += loc_helper.format_loc_text(LOC_DRAW_SINGLE, content=content)

            if item_selected.final_type == 1:  # 提前终止当前抽取
                feedback += loc_helper.format_loc_text(LOC_DRAW_FIN_INNER)
                break
        return feedback.strip()


@custom_user_command(readable_name="抽卡指令", priority=DPP_COMMAND_PRIORITY_DEFAULT)
class DeckCommand(UserCommandBase):
    """
    .draw 指令, 从牌库中抽取
    """

    def __init__(self, bot: Bot):
        super().__init__(bot)
        self.deck_dict: Dict[str, Deck] = {}

        bot.loc_helper.register_loc_text(LOC_DRAW_RESULT, "Draw {times} times from {deck_name}:\n{result}",
                                         f"抽卡回复, times为次数, deck_name为牌库名, result由{LOC_DRAW_SINGLE}和{LOC_DRAW_MULTI}定义")
        bot.loc_helper.register_loc_text(LOC_DRAW_RESULT_INLINE, "[Draw {times} times from {deck_name}:{result}]",
                                         f"嵌套抽卡内容, times为次数, deck_name为牌库名, "
                                         f"result由{LOC_DRAW_SINGLE}和{LOC_DRAW_MULTI}定义")
        bot.loc_helper.register_loc_text(LOC_DRAW_SINGLE, "{content}", "只抽一次时的回复的内容")
        bot.loc_helper.register_loc_text(LOC_DRAW_MULTI, "Result {time}: {content}", "抽取多次时单条内容, time为当前次数")
        bot.loc_helper.register_loc_text(LOC_DRAW_FIN_ALL, "Finalize draw! (All)", "抽到的内容使得所有抽取提前终止")
        bot.loc_helper.register_loc_text(LOC_DRAW_FIN_INNER, "Finalize draw! (Inner)", "抽到的内容使得内层抽取提前终止")
        bot.loc_helper.register_loc_text(LOC_DRAW_ERR_EMPTY_DECK, "Current decks is empty!", "牌库被抽光了(都是不放回的)")
        bot.loc_helper.register_loc_text(LOC_DRAW_ERR_TIME, "The draw time {times} is invalid!",
                                         "抽取次数不是合法正整数或不是合法的掷骰表达式, time为识别到的次数")
        bot.loc_helper.register_loc_text(LOC_DRAW_ERR_NO_DECK, "Cannot find deck {deck_name}", "找不到想要抽取的牌库")
        bot.loc_helper.register_loc_text(LOC_DRAW_ERR_VAGUE_DECK, "Possible decks: {deck_list}", "找到多个可能的牌库")

        bot.cfg_helper.register_config(CFG_DECK_DATA_PATH, "./DeckData", "牌库指令的数据来源, .代表Data文件夹")

    def delay_init(self) -> List[str]:
        # 从本地文件中读取资料
        data_path_list: List[str] = self.bot.cfg_helper.get_config(CFG_DECK_DATA_PATH)
        for i, path in enumerate(data_path_list):
            if path.startswith("./"):  # 用DATA_PATH作为当前路径
                data_path_list[i] = os.path.join(bot_config.DATA_PATH, path[2:])
        init_info: List[str] = []
        for data_path in data_path_list:
            self.load_data_from_path(data_path, init_info)
        for deck in self.deck_dict.values():
            for item in deck.items:
                try:
                    item.get_result(self.deck_dict.values(), self.bot.loc_helper, False)
                except ForceFinal:
                    pass
                except ValueError as e:
                    init_info.append(f"{deck.name}的第{deck.items.index(item)+1}个条目中存在错误: {e.args[0]}")
        init_info.append(self.get_state())
        return init_info

    def can_process_msg(self, msg_str: str, meta: MessageMetaData) -> Tuple[bool, bool, Any]:
        should_proc: bool = msg_str.startswith(".draw")
        should_pass: bool = False
        return should_proc, should_pass, msg_str[5:].strip()

    def process_msg(self, msg_str: str, meta: MessageMetaData, hint: Any) -> List[BotCommandBase]:
        port = GroupMessagePort(meta.group_id) if meta.group_id else PrivateMessagePort(meta.user_id)
        # 解析语句
        arg_str = hint
        feedback: str = ""
        args = arg_str.split("#", 1)
        if len(args) == 1:  # 没有给出Time
            times, deck_name = 1, args[0]
        else:  # 给定了Time
            times, deck_name = args[0], args[1]
            times = preprocess_roll_exp(times)
            if not is_roll_exp(times):
                feedback += self.format_loc(LOC_DRAW_ERR_TIME, times=times) + "\n"
            else:
                roll_res = exec_roll_exp(times)
                times = roll_res.get_val()
                if times <= 0 or times > DRAW_LIMIT:
                    feedback += self.format_loc(LOC_DRAW_ERR_TIME, times=times) + "\n"

        target_deck = None
        if deck_name in self.deck_dict.keys():
            target_deck = self.deck_dict[deck_name]
        else:
            poss_deck_names = match_substring(deck_name, self.deck_dict.keys())
            if len(poss_deck_names) == 0:
                feedback += self.format_loc(LOC_DRAW_ERR_NO_DECK, deck_name=deck_name) + "\n"
            elif len(poss_deck_names) > 1:
                poss_deck_names = [deck.name for deck in (self.deck_dict[name] for name in poss_deck_names)]
                feedback += self.format_loc(LOC_DRAW_ERR_VAGUE_DECK, deck_list=poss_deck_names) + "\n"
            else:
                deck_name = poss_deck_names[0]
                target_deck = self.deck_dict[deck_name]

        if target_deck:
            try:
                draw_result = target_deck.draw(times, self.deck_dict.values(), self.bot.loc_helper)
            except ForceFinal as e:
                draw_result = self.bot.loc_helper.format_loc_text(LOC_DRAW_SINGLE, content=e.info)
            feedback += self.format_loc(LOC_DRAW_RESULT, times=times, deck_name=target_deck.name, result=draw_result)
        feedback = feedback.strip()

        return [BotSendMsgCommand(self.bot.account, feedback, [port])]

    def get_help(self, keyword: str, meta: MessageMetaData) -> str:
        if keyword == "draw":  # help后的接着的内容
            feedback: str = ".draw [次数#, 可选][牌库名]" \
                            "示例: .draw 4#万象无常牌"
            return feedback
        return ""

    def get_description(self) -> str:
        return ".draw 抽卡指令"  # help指令中返回的内容

    def load_data_from_path(self, path: str, error_info: List[str]) -> None:
        """从指定文件或目录读取信息"""

        def load_data_from_xlsx(wb: openpyxl.Workbook):
            for sheet_name in wb.sheetnames:
                field_index_dict: Dict[str, int] = {}
                ws = wb[sheet_name]
                for header_cell in ws[1]:
                    if header_cell.value in DECK_ITEM_FIELD:
                        field_index_dict[header_cell.value] = header_cell.column - 1
                flag = False
                for field in DECK_ITEM_FIELD:
                    if field not in field_index_dict:
                        error_info.append(f"不完整的表格{path}/{sheet_name}, 缺少{field}, 未加载该工作表")
                        flag = True
                        break
                if flag:
                    continue

                # 生成Deck
                deck = Deck(sheet_name)
                # 逐行生成DeckItem
                row_index = 1
                for row in ws.iter_rows(min_row=2):
                    row_index += 1
                    content = row[field_index_dict[DECK_ITEM_FIELD_CONTENT]].value
                    content: str = str(content).strip() if content else ""

                    weight = row[field_index_dict[DECK_ITEM_FIELD_WEIGHT]].value
                    try:
                        weight = int(weight)
                        assert weight >= 1
                    except (TypeError, ValueError, AssertionError):
                        weight = 1

                    redraw = row[field_index_dict[DECK_ITEM_FIELD_REDRAW]].value
                    try:
                        redraw = int(redraw)
                        assert redraw in (0, 1)
                    except (TypeError, ValueError, AssertionError):
                        redraw = 1
                    redraw = True if redraw == 1 else False

                    final = row[field_index_dict[DECK_ITEM_FIELD_FINAL]].value
                    try:
                        final = int(final)
                        assert final in (0, 1, 2)
                    except (TypeError, ValueError, AssertionError):
                        final = 0

                    if not content:
                        dice_log(f"表格{sheet_name}第{row_index}行缺少content, 该条目未加载")
                        continue

                    item = DeckItem(content, weight, redraw, final)
                    deck.add_item(item)

                # 记录到self字典中
                if deck.items:
                    deck_name = preprocess_msg(deck.name)  # 预处理一下名字, 防止输入的大小写被预处理后无法匹配
                    self.deck_dict[deck_name] = deck

        if path.endswith(".xlsx"):
            if os.path.exists(path):
                try:
                    workbook = read_xlsx(path)
                except PermissionError:
                    error_info.append(f"读取{path}时遇到错误: 权限不足")
                    return
                load_data_from_xlsx(workbook)
            else:
                create_parent_dir(path)  # 父文件夹不存在需先创建父文件夹
                workbook = get_template_query_workbook()
                update_xlsx(workbook, path)
                workbook.close()
        elif path:  # 是文件夹
            if os.path.exists(path):
                try:
                    inner_paths = os.listdir(path)  # 遍历文件夹下所有文件
                    for inner_path in inner_paths:
                        inner_path = os.path.join(path, inner_path)
                        self.load_data_from_path(inner_path, error_info)
                except FileNotFoundError as e:  # 文件夹不存在
                    error_info.append(f"读取{path}时遇到错误: {e}")
            else:
                create_parent_dir(path)

    def get_state(self) -> str:
        feedback: str
        if self.deck_dict:
            feedback = f"已加载{len(self.deck_dict)}个牌库: {[deck.name for deck in self.deck_dict.values()]}"
        else:
            feedback = f"尚未加载任何牌库"
        return feedback


def get_template_query_workbook() -> openpyxl.Workbook:
    """获得一个模板工作簿"""
    wb = openpyxl.Workbook()
    for name in wb.sheetnames:
        del wb[name]
    ws_temp = wb.create_sheet("template")
    for i, text in enumerate(DECK_ITEM_FIELD):
        cell_field = ws_temp.cell(row=1, column=1 + i, value=text)
        cell_field.comment = Comment(DECK_ITEM_FIELD_COMMENT[text], "DicePP")
    return wb
