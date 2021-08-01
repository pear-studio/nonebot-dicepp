from typing import List, Tuple, Dict, Optional, Set, Literal, Iterable, Union, Any
import os
import datetime
import openpyxl
from openpyxl.comments import Comment
import random
from openpyxl.utils import get_column_letter

import bot_utils
from bot_core import Bot
import bot_config
from command.command_config import *
from command.dicepp_command import UserCommandBase, custom_user_command, MessageMetaData
from command.bot_command import BotCommandBase, MessagePort, PrivateMessagePort, GroupMessagePort, BotSendMsgCommand
from data_manager import DataManagerError
from bot_utils.localdata import read_xlsx, update_xlsx, create_parent_dir

# LOC_NICKNAME_SET = "nickname_set"

CFG_QUERY_DATA_PATH = "query_data_path"

LOC_QUERY_RESULT = "query_result"
LOC_QUERY_SINGLE_RESULT = "query_single_result"
LOC_QUERY_MULTI_RESULT = "query_multi_result"
LOC_QUERY_MULTI_RESULT_ITEM = "query_multi_result_item"
LOC_QUERY_MULTI_RESULT_PAGE = "query_multi_result_page"
LOC_QUERY_MULTI_RESULT_PAGE_UNDERFLOW = "query_multi_result_page_underflow"
LOC_QUERY_MULTI_RESULT_PAGE_OVERFLOW = "query_multi_result_page_overflow"
LOC_QUERY_NO_RESULT = "query_no_result"
LOC_QUERY_KEY_NUM_EXCEED = "query_key_num_exceed"

QUERY_ITEM_FIELD_KEY = "Key"
QUERY_ITEM_FIELD_SYN = "Synonym"
QUERY_ITEM_FIELD_CONTENT = "Content"
QUERY_ITEM_FIELD_DESC = "Description"
QUERY_ITEM_FIELD_CAT = "Catalogue"
QUERY_ITEM_FIELD_TAG = "Tag"

QUERY_ITEM_FIELD = [QUERY_ITEM_FIELD_KEY, QUERY_ITEM_FIELD_SYN, QUERY_ITEM_FIELD_CONTENT, QUERY_ITEM_FIELD_DESC,
                    QUERY_ITEM_FIELD_CAT, QUERY_ITEM_FIELD_TAG]
QUERY_ITEM_FIELD_COMMENT = {QUERY_ITEM_FIELD_KEY: "查询关键字",
                            QUERY_ITEM_FIELD_SYN: "选填, 关键字同义词, 用/分割",
                            QUERY_ITEM_FIELD_CONTENT: "词条内容",
                            QUERY_ITEM_FIELD_DESC: "选填, 对于词条的简短描述",
                            QUERY_ITEM_FIELD_CAT: "选填, 该词条被归在哪一个目录下, 用/分割父子目录",
                            QUERY_ITEM_FIELD_TAG: "选填, 内容的Tag, 如#核心 #法术",
                            }

QUERY_ITEM_FIELD_DESC_DEFAULT_LEN = 20  # 默认用前多少个字符作为默认Description

MAX_QUERY_KEY_NUM = 5  # 最多能同时用多少个查询关键字
MAX_QUERY_CANDIDATE_NUM = 10  # 详细查询时一页最多能同时展示多少个条目
RECORD_RESPONSE_TIME = 60  # 至多响应多久以前的查询指令, 多余的将被清理
RECORD_CLEAN_FREQ = 50  # 每隔多少次查询指令尝试清理一次查询记录


class QueryItem:
    def __init__(self, uuid: int, key: Tuple[str], content: str, desc: str, catalogue: str, tag: List[str]
                 , src_file_uuid: int, src_file_index: int):
        """表示一个可以被查询到的条目"""
        self.uuid = uuid  # 唯一标识符
        self.key = key  # 查询关键字
        self.content = content  # 条目内容
        self.desc = desc  # 条目描述
        self.catalogue = catalogue  # 出自哪一个目录, 如"玩家手册/第一章"
        self.tag = tag  # 条目的Tag列表, 如[战士, 天赋]

        self.src_file_uuid = src_file_uuid  # 用来标志这个条目是从哪个源文件读取的, uuid对应的path从Command实例中找到
        self.src_file_row = src_file_index  # 属于源文件的第几个条目, 从1开始


class QuerySource:
    def __init__(self, uuid: int, path: str, sheet: str):
        """表示一个可以被查询到的条目"""
        self.uuid = uuid  # 唯一标识符
        self.path = path  # 文件路径
        self.sheet = sheet  # 分页名称


class QueryRecord:
    def __init__(self, uuid_list: List[int], time: datetime.datetime):
        """记录一次可交互的查询指令"""
        self.uuid_list = uuid_list  # 当前可以选择的条目的uuid列表
        self.time = time  # 更新时间
        self.page = 1  # 当前的页数


@custom_user_command(priority=2,
                     group_only=False,
                     flag=DPP_COMMAND_FLAG_DEFAULT,
                     cluster=DPP_COMMAND_CLUSTER_DEFAULT)
class QueryCommand(UserCommandBase):
    """
    查询资料库的指令, 以.查询或.q开头
    """

    def __init__(self, bot: Bot):
        super().__init__(bot)
        self.query_dict: Dict[str, List[int]] = {}  # key为查询关键字, value为item uuid
        self.item_uuid_dict: Dict[int, QueryItem] = {}  # key为item uuid
        self.src_uuid_dict: Dict[int, QuerySource] = {}  # key为source uuid
        self.record_dict: Dict[MessagePort, QueryRecord] = {}
        self.record_clean_flag: int = 0

        reg_loc = bot.loc_helper.register_loc_text
        reg_loc(LOC_QUERY_RESULT, "{result}", "查询成功时返回的内容")
        reg_loc(LOC_QUERY_SINGLE_RESULT, "{keyword}: {tag}\n{content}{cat}{syn}",
                "查询找到唯一条目, keyword: 主关键字, content: 词条内容"
                ", cat: 换行+目录*, tag: 标签*, syn: 换行+同义词*  (*如果有则显示, 没有则忽略)")
        reg_loc(LOC_QUERY_MULTI_RESULT, "{multi_results}",
                f"查询找到多个条目, multi_results由多行{LOC_QUERY_MULTI_RESULT_ITEM}构成")
        reg_loc(LOC_QUERY_MULTI_RESULT_ITEM, "{keyword}: {description} {cat} {tag} {syn}",
                "查询找到多个条目时单个条目的描述, keyword: 主关键字, description: 简短描述"
                ", cat: 目录*, tag: 标签*, syn: 同义词*  (*如果有则显示, 没有则忽略)")
        reg_loc(LOC_QUERY_MULTI_RESULT_PAGE, "Page{page_cur}/{page_total}, + for next page, - for prev page",
                "搜索结果出现多页时提示, {page_cur}表示当前页, {page_total}表示总页数")
        reg_loc(LOC_QUERY_MULTI_RESULT_PAGE_UNDERFLOW, "This is the first page!", "用户尝试在第一页往前翻页时的提醒")
        reg_loc(LOC_QUERY_MULTI_RESULT_PAGE_OVERFLOW, "This is the final page!", "用户尝试在最后一页往后翻页时的提醒")
        reg_loc(LOC_QUERY_NO_RESULT, "Cannot find result...", "查询失败时的提示")
        reg_loc(LOC_QUERY_KEY_NUM_EXCEED, "Maximum keyword num is {key_num}",
                "用户查询时使用过多关键字时的提示 {key_num}为关键字数量上限")

        bot.cfg_helper.register_config(CFG_QUERY_DATA_PATH, "./QueryData", "查询指令的数据来源, .代表Data文件夹")

    def delay_init(self) -> List[str]:
        # 从本地文件中读取资料
        data_path_list: List[str] = self.bot.cfg_helper.get_config(CFG_QUERY_DATA_PATH)
        for i, path in enumerate(data_path_list):
            if path[:2] == "./":  # 用DATA_PATH作为当前路径
                data_path_list[i] = os.path.join(bot_config.DATA_PATH, path[2:])
        error_info: List[str] = []
        for data_path in data_path_list:
            self.load_data_from_path(data_path, error_info)
        return error_info

    def can_process_msg(self, msg_str: str, meta: MessageMetaData) -> Tuple[bool, bool, Any]:
        should_proc: bool = False
        should_pass: bool = False
        mode: Optional[Literal["query", "search", "select", "flip_page"]] = None
        arg_str: Optional[str] = None

        # 响应交互查询指令
        port = MessagePort(meta.group_id, meta.user_id)
        if port in self.record_dict:
            record = self.record_dict[port]
            if bot_utils.time.get_current_date_raw() - record.time < datetime.timedelta(seconds=RECORD_RESPONSE_TIME):
                # 选择条目
                try:
                    should_proc = (0 < int(msg_str) <= len(record.uuid_list))
                    mode, arg_str = "select", msg_str
                except ValueError:
                    pass
                # 翻页
                if not should_proc and msg_str.strip() == "+":
                    should_proc, mode, arg_str = True, "flip_page", "+"
                if not should_proc and msg_str.strip() == "-":
                    should_proc, mode, arg_str = True, "flip_page", "-"
            else:
                del self.record_dict[port]  # 清理过期条目

        # 常规查询指令
        for key in ["查询", "q"]:
            if not should_proc and msg_str.startswith(f".{key}"):
                should_proc, mode, arg_str = True, "query", msg_str[1 + len(key):].strip()
        for key in ["搜索", "s"]:
            if not should_proc and msg_str.startswith(f".{key}"):
                should_proc, mode, arg_str = True, "search", msg_str[1 + len(key):].strip()

        # ToDo: 在线修改本地查询条目

        assert (not should_proc) or mode
        hint = (mode, arg_str)
        return should_proc, should_pass, hint

    def process_msg(self, msg_str: str, meta: MessageMetaData, hint: Any) -> List[BotCommandBase]:
        port = GroupMessagePort(meta.group_id) if meta.group_id else PrivateMessagePort(meta.user_id)
        source_port = MessagePort(meta.group_id, meta.user_id)
        mode: Literal["query", "search", "select", "flip_page"] = hint[0]
        arg_str: str = hint[1]
        feedback: str

        # 处理指令
        if mode == "query":
            feedback = self.query_info(arg_str, source_port, search_mode=0)
        elif mode == "search":
            feedback = self.query_info(arg_str, source_port, search_mode=1)
        elif mode == "select":
            index = int(arg_str)
            uuid = self.record_dict[source_port].uuid_list[index]
            item: QueryItem = self.item_uuid_dict[uuid]
            feedback = self.format_single_item_feedback(item)
        elif mode == "flip_page":
            record = self.record_dict[source_port]
            next_page = (arg_str == "+")
            feedback, cur_page = self.flip_page(record, next_page)
            self.record_dict[source_port].page = cur_page
        else:
            raise NotImplementedError()

        # 尝试清理过期的查询记录
        self.record_clean_flag += 1
        if self.record_clean_flag >= RECORD_CLEAN_FREQ:
            self.record_clean_flag = 0
            self.clean_records()

        return [BotSendMsgCommand(self.bot.account, feedback, [port])]

    def get_help(self, keyword: str, meta: MessageMetaData) -> str:
        if keyword in ["查询", "索引", "q", "i"]:
            help_str = "设置昵称：.nn [昵称]\n" \
                       "私聊.nn视为操作全局昵称\n" \
                       "昵称优先级:群昵称>私聊昵称>群名片>QQ昵称\n" \
                       "群聊中的nn指令会智能修改先攻列表中的名字\n" \
                       "示例:\n" \
                       ".nn	//视为删除昵称\n" \
                       ".nn dm //将昵称设置为dm"
            return help_str
        return ""

    def get_description(self) -> str:
        return ".查询 根据关键字查找资料 .索引 根据资料内容查找资料"

    def query_info(self, query_key: str, port: MessagePort, search_mode: int) -> str:
        """
        查询信息, 返回输出给用户的字符串, 若给出选项将会记录信息以便响应用户之后的快速查询.
        search_mode != 0则使用模糊查找
        """
        # 清空过往记录
        if port in self.record_dict:
            del self.record_dict[port]
        # 分割关键字
        query_key_list = [key.strip() for key in query_key.split("/") if key.strip()]
        if len(query_key_list) > MAX_QUERY_KEY_NUM:  # 关键字数量超出上限
            return self.format_loc(LOC_QUERY_KEY_NUM_EXCEED, key_num=MAX_QUERY_KEY_NUM)
        # 找到搜索候选
        poss_result = self.search_item(query_key_list, search_mode)
        # 处理
        if not poss_result:  # 找不到结果
            return self.format_loc(LOC_QUERY_NO_RESULT)
        elif len(poss_result) == 1:  # 找到唯一结果
            item: QueryItem = self.item_uuid_dict[poss_result[0]]
            return self.format_single_item_feedback(item)
        else:  # len(poss_result) > 1  找到多个结果, 记录当前信息并提示用户选择
            # 记录当前信息以备将来查询
            self.record_dict[port] = QueryRecord(poss_result, bot_utils.time.get_current_date_raw())
            show_result: List[int] = poss_result[:MAX_QUERY_CANDIDATE_NUM]
            items: List[QueryItem] = [self.item_uuid_dict[uuid] for uuid in show_result]
            feedback = self.format_multiple_items_feedback(items)
            if len(poss_result) > MAX_QUERY_CANDIDATE_NUM:
                feedback += "\n" + self.format_loc(LOC_QUERY_MULTI_RESULT_PAGE, page_cur=1,
                                                   page_total=len(poss_result) // MAX_QUERY_CANDIDATE_NUM + 1)
            return feedback

    def search_item(self, query_key_list: List[str], search_mode: int) -> List[int]:
        poss_result: List[int] = []
        if search_mode == 0:  # 仅匹配条目关键字
            for candidate_key in self.query_dict.keys():
                if all(((key in candidate_key) for key in query_key_list)):  # 所有关键字都在candidate_key中出现
                    poss_result += self.query_dict[candidate_key]
        else:  # 匹配条目关键字和内容
            for item in self.item_uuid_dict.values():
                candidate_key: str = "/".join(list(item.key) + [item.content])
                if all(((key in candidate_key) for key in query_key_list)):  # 所有关键字都在candidate_key中出现
                    poss_result.append(item.uuid)  # self.query_dict[item.key[0]]
        # 去除重复的条目
        poss_result = list(set(poss_result))
        return poss_result

    def format_single_item_feedback(self, item: QueryItem) -> str:
        item_keyword = item.key[0]
        item_content = item.content
        item_tag = get_tag_string(item.tag)
        item_cat = "\n目录: " + item.catalogue if item.catalogue else ""
        item_syn = "\n同义词: " + get_syn_string(item.key[1:]) if item.key[1:] else ""
        return self.format_loc(LOC_QUERY_SINGLE_RESULT, keyword=item_keyword, content=item_content, tag=item_tag,
                               cat=item_cat, syn=item_syn)

    def format_multiple_items_feedback(self, items: List[QueryItem]):
        feedback = ""
        for index, item in enumerate(items):
            item_keyword = item.key[0]
            item_desc = item.desc
            item_tag = " Tag: " + get_tag_string(item.tag) if item.catalogue else ""
            item_cat = " 目录: " + item.catalogue if item.catalogue else ""
            item_syn = " 同义词: " + get_syn_string(item.key[1:]) if item.key[1:] else ""
            item_info = self.format_loc(LOC_QUERY_MULTI_RESULT_ITEM, keyword=item_keyword, description=item_desc,
                                        tag=item_tag, cat=item_cat, syn=item_syn)
            feedback += f"{index}. {item_info}\n"
        feedback = feedback.strip()
        return feedback

    def flip_page(self, record: QueryRecord, next_page: bool) -> Tuple[str, int]:
        cur_page = record.page
        total_page = len(record.uuid_list) // MAX_QUERY_CANDIDATE_NUM + 1
        if not next_page:
            if cur_page == 1:
                feedback = self.format_loc(LOC_QUERY_MULTI_RESULT_PAGE_UNDERFLOW)
            else:
                cur_page = cur_page - 1
                index = (cur_page - 1) * MAX_QUERY_CANDIDATE_NUM
                uuids = record.uuid_list[index:index + MAX_QUERY_CANDIDATE_NUM]
                items = [self.item_uuid_dict[uuid] for uuid in uuids]
                feedback = self.format_multiple_items_feedback(items)
        else:
            if cur_page == total_page:
                feedback = self.format_loc(LOC_QUERY_MULTI_RESULT_PAGE_OVERFLOW)
            else:
                cur_page = cur_page + 1
                index = (cur_page - 1) * MAX_QUERY_CANDIDATE_NUM
                uuids = record.uuid_list[index:index + MAX_QUERY_CANDIDATE_NUM]
                items = [self.item_uuid_dict[uuid] for uuid in uuids]
                feedback = self.format_multiple_items_feedback(items)
        if len(record.uuid_list) > MAX_QUERY_CANDIDATE_NUM:
            feedback += "\n" + self.format_loc(LOC_QUERY_MULTI_RESULT_PAGE, page_cur=cur_page, page_total=total_page)
        return feedback, cur_page

    def clean_records(self):
        """清理过期的查询指令"""
        invalid_ports: Set[MessagePort] = set()
        for port, record in self.record_dict.items():
            if bot_utils.time.get_current_date_raw() - record.time > datetime.timedelta(seconds=RECORD_RESPONSE_TIME):
                invalid_ports.add(port)
        for port in invalid_ports:
            del self.record_dict[port]

    def load_data_from_path(self, path: str, error_info: List[str]) -> None:
        """从指定文件或目录读取信息"""

        def load_data_from_xlsx(wb: openpyxl.Workbook):
            for sheet_name in wb.sheetnames:
                field_index_dict: Dict[str, str] = {}
                ws = wb[sheet_name]
                for header_cell in ws[1]:
                    if header_cell.value in QUERY_ITEM_FIELD:
                        field_index_dict[header_cell.value] = header_cell.column - 1
                flag = False
                for field in QUERY_ITEM_FIELD:
                    if field not in field_index_dict:
                        error_info.append(f"不完整的表格{path}/{sheet_name}, 缺少{field}, 未加载该工作表")
                        flag = True
                if flag:
                    continue

                # 生成QuerySource
                qs_uuid = get_query_uuid()
                qs = QuerySource(qs_uuid, path, sheet_name)
                # 逐行生成QueryItem
                qi_list: List[QueryItem] = []
                row_index = 1
                for row in ws.iter_rows(min_row=2):
                    row_index += 1
                    main_key = row[field_index_dict[QUERY_ITEM_FIELD_KEY]].value
                    main_key: str = str(main_key).strip() if main_key else ""

                    item_key = row[field_index_dict[QUERY_ITEM_FIELD_SYN]].value
                    item_key = str(item_key).strip().split("/") if item_key else []  # 用/分隔同义词
                    item_key: List[str] = [main_key] + [syn.strip() for syn in item_key if syn.strip()]
                    item_key_tuple: Tuple[str] = tuple(item_key)

                    item_content = row[field_index_dict[QUERY_ITEM_FIELD_CONTENT]].value
                    item_content: str = str(item_content).strip() if item_content else ""

                    item_desc = row[field_index_dict[QUERY_ITEM_FIELD_DESC]].value
                    item_desc: str = str(item_desc).strip() if item_desc else ""

                    item_cat = row[field_index_dict[QUERY_ITEM_FIELD_CAT]].value
                    item_cat: str = str(item_cat).strip() if item_cat else ""

                    item_tag = row[field_index_dict[QUERY_ITEM_FIELD_TAG]].value
                    item_tag = str(item_tag).strip().split("#") if item_tag else []  # 用#分隔Tag
                    item_tag: List[str] = [tag.strip() for tag in item_tag if tag.strip()]

                    if not main_key:
                        error_info.append(f"表格{wb.path}/{sheet_name}第{row_index}缺少key, 该条目未加载")
                        continue
                    if not item_content:
                        error_info.append(f"表格{wb.path}/{sheet_name}第{row_index}缺少content, 该条目未加载")
                        continue
                    if not item_desc:  # 用content的前一部分自动生成desc
                        item_desc = item_content[:QUERY_ITEM_FIELD_DESC_DEFAULT_LEN].replace("\n", " ") + "..."
                    if not item_tag:
                        item_tag = []

                    qi_uuid = get_query_uuid()
                    qi = QueryItem(qi_uuid, item_key_tuple, item_content, item_desc, item_cat, item_tag,
                                   qs_uuid, row_index - 1)
                    qi_list.append(qi)
                # 记录到self字典中
                if qi_list:
                    self.src_uuid_dict[qs_uuid] = qs
                    for item in qi_list:
                        self.item_uuid_dict[item.uuid] = item
                        for k in item.key:
                            if k not in self.query_dict:
                                self.query_dict[k] = [item.uuid]
                            else:
                                self.query_dict[k].append(item.uuid)

        if path.endswith(".xlsx"):
            if os.path.exists(path):
                workbook = read_xlsx(path)
                load_data_from_xlsx(workbook)
            else:
                create_parent_dir(path)  # 父文件夹不存在需先创建父文件夹
                workbook = get_template_query_workbook()
                update_xlsx(workbook, path)
                workbook.close()
        elif "." not in path:  # 是文件夹
            try:
                inner_paths = os.listdir(path)  # 遍历文件夹下所有文件
                for inner_path in inner_paths:
                    inner_path = os.path.join(path, inner_path)
                    self.load_data_from_path(inner_path, error_info)
            except FileNotFoundError:  # 文件夹不存在
                pass


def get_template_query_workbook() -> openpyxl.Workbook:
    """获得一个模板工作簿"""
    wb = openpyxl.Workbook()
    for name in wb.sheetnames:
        del wb[name]
    ws_temp = wb.create_sheet("template")
    for i, text in enumerate(QUERY_ITEM_FIELD):
        cell_field = ws_temp.cell(row=1, column=1 + i, value=text)
        cell_field.comment = Comment(QUERY_ITEM_FIELD_COMMENT[text], "DicePP")
    return wb


QUERY_UUID_SET: Set[int] = set()
QUERY_UUID_MIN: int = -(1 << 31)
QUERY_UUID_MAX: int = 1 << 31


def get_query_uuid() -> int:
    iter_num = 0
    uuid = random.randint(QUERY_UUID_MIN, QUERY_UUID_MAX)
    while uuid in QUERY_UUID_SET:
        uuid = random.randint(QUERY_UUID_MIN, QUERY_UUID_MAX)
        iter_num += 1
        if iter_num > 100000:
            raise TimeoutError("获取Query UUID失败!")
    QUERY_UUID_SET.add(uuid)
    return uuid


def get_tag_string(tags: Iterable[str]):
    """获得标签字符串, 输入["a", "b"], 返回"#a #b" """
    return " ".join(["#" + tag for tag in tags])


def get_syn_string(synonym: Iterable[str]):
    """获得同义词字符串, ["a", "b"], 返回"a, b" """
    return ", ".join(synonym)
