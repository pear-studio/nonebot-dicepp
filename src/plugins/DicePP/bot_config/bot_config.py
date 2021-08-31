import os
from typing import Dict, List

import openpyxl
from openpyxl.comments import Comment

import logger
from bot_utils.localdata import read_xlsx

from bot_config.config_item import ConfigItem

PROJECT_PATH = os.path.dirname(os.path.dirname(__file__))

DATA_PATH = os.path.join(PROJECT_PATH, 'Data')

BOT_DATA_PATH = os.path.join(DATA_PATH, 'Bot')
CONFIG_PATH = os.path.join(DATA_PATH, 'Config')
LOCAL_IMG_PATH = os.path.join(CONFIG_PATH, 'LocalImage')

ALL_LOCAL_DIR_PATH = [DATA_PATH, BOT_DATA_PATH, CONFIG_PATH, LOCAL_IMG_PATH]

for dirPath in ALL_LOCAL_DIR_PATH:
    if not os.path.exists(dirPath):
        os.makedirs(dirPath)
        logger.dice_log("[Config] [Init] 创建文件夹: " + dirPath)

BOT_VERSION = "Ver 1.0.0 (Early Access 210826)"

BOT_DESCRIBE = "DicePP by 梨子"

BOT_AGREEMENT = "1.邀请骰娘, 使用掷骰服务和在群内阅读此协议视为同意并承诺遵守此协议，否则请移除骰娘。\n" \
                "2.不允许禁言骰娘或刷屏掷骰等对骰娘的不友善行为，这些行为将会提高骰娘被制裁的风险。开关骰娘响应请使用.bot on/off。\n" \
                "3.邀请骰娘入群应已事先得到群内同意。因擅自邀请而使骰娘遭遇不友善行为时，邀请者因未履行预见义务而将承担连带责任。\n" \
                "4.禁止将骰娘用于赌博及其他违法犯罪行为，禁止将本骰娘用作TRPG外的用途，禁止拉入非TRPG群。\n" \
                "5.对于设置敏感昵称等无法预见但有可能招致言论审查的行为，骰娘可能会出于自我保护而拒绝提供服务\n" \
                "6.由于技术以及资金原因，无法保证骰娘100%的时间稳定运行，可能不定时停机维护或遭遇冻结，敬请谅解。\n" \
                "7.对于违反协议的行为，骰娘将视情况终止对用户和所在群提供服务。\n" \
                "8.本协议内容可能改动，请注意查看最新协议。\n" \
                "9.本服务最终解释权归服务提供方所有。"

BOT_GIT_LINK = "https://github.com/pear-studio/nonebot-dicepp"

DEFAULT_CONFIG: Dict[str, str] = {}
DEFAULT_CONFIG_COMMENT: Dict[str, str] = {}

# 默认配置
CFG_MASTER = "master"
DEFAULT_CONFIG[CFG_MASTER] = ""
DEFAULT_CONFIG_COMMENT[CFG_MASTER] = "Master账号, 权限最高, 可以有多个Master"

CFG_ADMIN = "admin"
DEFAULT_CONFIG[CFG_ADMIN] = ""
DEFAULT_CONFIG_COMMENT[CFG_ADMIN] = "管理员账号, 拥有次高权限, 可以有多个管理员"

CFG_FRIEND_TOKEN = "friend_token"
DEFAULT_CONFIG[CFG_FRIEND_TOKEN] = ""
DEFAULT_CONFIG_COMMENT[CFG_FRIEND_TOKEN] = "用户申请好友时在验证中输入参数中的文本之一骰娘才会通过, 若字符串为空则通过所有的好友验证"

CFG_AGREEMENT = "agreement"
DEFAULT_CONFIG[CFG_AGREEMENT] = BOT_AGREEMENT
DEFAULT_CONFIG_COMMENT[CFG_AGREEMENT] = "使用协议"

CONFIG_FILE_PATH = "config.xlsx"

CFG_COMMAND_SPLIT = "command_split"  # \\ 来分割多条指令
DEFAULT_CONFIG[CFG_COMMAND_SPLIT] = "\\\\"
DEFAULT_CONFIG_COMMENT[CFG_COMMAND_SPLIT] = "分割多条指令的关键字, 默认为 \\\\"


class ConfigHelper:
    def __init__(self, data_path: str, identifier: str):
        self.data_path = os.path.join(data_path, CONFIG_FILE_PATH)
        self.identifier = identifier
        self.all_configs: Dict[str, ConfigItem] = {}

        # 默认配置
        for key in DEFAULT_CONFIG.keys():
            self.register_config(key, DEFAULT_CONFIG[key], DEFAULT_CONFIG_COMMENT[key])

    def load_config(self):
        """用文档里的配置覆写之前的配置"""
        if not os.path.exists(self.data_path):
            return
        workbook = read_xlsx(self.data_path)
        if self.identifier in workbook.sheetnames:
            sheet = workbook[self.identifier]
        elif "Default" in workbook.sheetnames:
            sheet = workbook["Default"]
        else:
            return
        for row in sheet.iter_rows():
            key = str(row[0].value)  # 第一个元素为关键字
            if key not in self.all_configs:
                continue
            comment: str = self.all_configs[key].comment  # 沿用原来的注释, 不用文件里的
            self.all_configs[key] = ConfigItem(key, comment=comment)
            for text in [str(cell.value) for cell in row[1:] if cell.value]:
                self.all_configs[key].add(text)
        workbook.close()
        logger.dice_log(f"[BotConfig] [Load] 成功读取配置文件 {self.data_path}")

    def save_config(self):
        """按现在的设置多个机器人会读写同一个配置文件, 如果并行可能存在写冲突, 现在应该是单线程异步, 应该没问题"""
        def save_loc_text_to_row(sheet, item: ConfigItem, row: int):
            header = sheet.cell(row=row, column=1, value=item.key)
            if item.comment:
                header.comment = Comment(item.comment, "DicePP")
            if not item.contents:
                sheet.cell(row=row, column=2, value="")
            else:
                for ci, text in enumerate(item.contents):
                    sheet.cell(row=row, column=ci + 2, value=text)

        feedback: str
        if os.path.exists(self.data_path):
            workbook = read_xlsx(self.data_path)
            feedback = "成功更新配置文件"
        else:
            workbook = openpyxl.Workbook()
            for name in workbook.sheetnames:
                del workbook[name]
            feedback = "成功创建配置文件"
        if self.identifier in workbook.sheetnames:
            cur_sheet = workbook[self.identifier]
        else:
            cur_sheet = workbook.create_sheet(self.identifier)
        for ri, cfg_item in enumerate(self.all_configs.values()):
            save_loc_text_to_row(cur_sheet, cfg_item, ri + 1)

        try:
            workbook.save(self.data_path)
        except PermissionError:
            logger.dice_log(f"[BotConfig] [Save] Save config {self.data_path} failed as permission denied!")

        workbook.close()

        logger.dice_log(f"[BotConfig] [Save] {feedback} {self.data_path}")

    def register_config(self, key: str, origin_str: str, comment: str = ""):
        """
        将一个配置注册至Helper中
        Args:
            key: 配置关键字
            origin_str: 配置的默认值
            comment: 注释
        """
        self.all_configs[key] = ConfigItem(key, origin_str, comment)

    def get_config(self, key: str) -> List[str]:
        """
        获取本地化语句
        Args:
            key: 本地化语句的关键字
        """
        return self.all_configs[key].get()
