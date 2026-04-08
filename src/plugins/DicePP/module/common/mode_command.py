from typing import Dict, Optional, List, Tuple, Any
import openpyxl
import os
import re

from core.bot import Bot
from core.data.models.extended import GroupConfig
from core.command.const import *
from core.command import UserCommandBase, custom_user_command
from core.command import BotCommandBase, BotSendMsgCommand
from core.command import CommandTextParser, CommandContextResolver
from core.communication import MessageMetaData, PrivateMessagePort, GroupMessagePort
from core.localization import LOC_PERMISSION_DENIED_NOTICE, LOC_FUNC_DISABLE

# Task 3.3: 统一解析器（替代内嵌前缀判断与参数切分）
_MODE_PARSER_ZH = CommandTextParser(command_prefix="模式", strip_prefix_len=3)
_MODE_PARSER_EN = CommandTextParser(command_prefix="mode", strip_prefix_len=5)

LOC_MODE_SWITCH = "mode_switch"
LOC_MODE_INVALID = "mode_invalid"
LOC_MODE_NOT_EXIST = "mode_not_exist"
LOC_MODE_LIST = "mode_list"
LOC_MODE_LIKELY = "mode_likely"
LOC_MODE_CURRENT = "mode_current"
LOC_MODE_DB_MATCH = "mode_db_match"
LOC_MODE_DB_MULTI_MATCH = "mode_db_multi_match"

CFG_MODE_ENABLE = "mode_enable"
CFG_MODE_DEFAULT = "mode_default"

MODE_FILE_PATH = "Config/mode_setting.xlsx"

DEFAULT_FIELD = ['mode', 'default_dice', 'query_database']
DEFAULT_TABLE = [
    ["DND5E2024", "20", "DND5E2024"],
    ["DND5E2014", "20", "DND5E2014"],
    ["DND5E混用", "20", "DND5E混合"],
    ["DND3R", "20", "DND3R"],
    ["PF1E", "20", "PF1E"],
    ["PF2E", "20", "PF2E"],
    ["PF2R", "20", "PF2R"],
    ["COC7", "100", "COC7"],
    ["NECHRONICA", "10", "NECHRONICA"],
]


@custom_user_command(readable_name="模式指令", priority=-2,
                     flag=DPP_COMMAND_FLAG_MANAGE, group_only=False
                     )
class ModeCommand(UserCommandBase):
    """
    .mode 模式设置指令（批量群设置修改/模板调用指令）
    """

    def __init__(self, bot: Bot):
        super().__init__(bot)
        bot.loc_helper.register_loc_text(LOC_MODE_SWITCH, "已切换至{new_mode}模式（默认{dice}面骰点，查询数据库使用{database}.db（如果有））。",
                                         "。mode切换群模式指令，切换模式等于一次性修改多个群配置。\nnew_mode：切换后的模式，dice：默认骰面，database：查询使用数据库")
        bot.loc_helper.register_loc_text(
            LOC_MODE_INVALID, "该模式配置有误，无法切换，请询问骰主。", "。mode切换群模式，但模式文件有问题的情况下返回")
        bot.loc_helper.register_loc_text(
            LOC_MODE_NOT_EXIST, "该模式不存在！", "。mode切换群模式，但模式文件中不存在此模式时返回")
        bot.loc_helper.register_loc_text(
            LOC_MODE_LIST, "以下是可用的模式列表：{modes}", "。mode模式指令查看可用模式列表\nmodes：可用模式列表")
        bot.loc_helper.register_loc_text(
            LOC_MODE_LIKELY, "找到多个选项，你要找的是不是：{modes}", "。mode模式指令，模糊匹配出现多个结果\nmodes：模糊匹配结果列表")
        bot.loc_helper.register_loc_text(LOC_MODE_CURRENT, "当前模式为{new_mode}（默认{dice}面骰点，查询数据库使用{database}.db（如果有））。", ".mode 不带参数时显示当前模式")
        bot.loc_helper.register_loc_text(LOC_MODE_DB_MATCH, "已自动匹配到数据库{database}（默认{dice}面骰点）。", ".mode自动匹配数据库时返回")
        bot.loc_helper.register_loc_text(LOC_MODE_DB_MULTI_MATCH, "找到多个匹配的数据库：{databases}，请使用更精确的名称。", ".mode自动匹配到多个数据库时返回")


        self.mode_dict: Dict[str, List[str]] = {}
        self.mode_field: List[str] = DEFAULT_FIELD
        # 大写模式名到原始模式名映射，减少重复遍历
        self.mode_upper_map: Dict[str, str] = {}

    def delay_init(self) -> List[str]:
        bot_id: str = self.bot.account
        init_info: List[str] = []
        edited: bool = False
        # 从本地文件中读取可用模式一览
        data_path = os.path.join(self.bot.data_path, MODE_FILE_PATH)
        os.makedirs(os.path.dirname(data_path), exist_ok=True)
        if os.path.exists(data_path):
            wb = openpyxl.load_workbook(data_path)
            id_list = wb.sheetnames
            if bot_id in id_list:
                ws = wb[bot_id]
                for row in ws:
                    if str(row[0].value) == "mode":
                        self.mode_field = [str(cell.value) for cell in row]
                    else:
                        self.mode_dict[str(row[0].value)] = [
                            str(cell.value) for cell in row[1:]]
            else:
                ws = wb.create_sheet(bot_id)
                ws.append(DEFAULT_FIELD)
                for str_row in DEFAULT_TABLE:
                    ws.append(str_row)
                    self.mode_dict[str_row[0]] = [
                        str_cell for str_cell in str_row[1:]]
                edited = True
            init_info.append("已载入模式文件。")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = bot_id
            ws.append(DEFAULT_FIELD)
            for str_row in DEFAULT_TABLE:
                ws.append(str_row)
                self.mode_dict[str_row[0]] = [
                    str_cell for str_cell in str_row[1:]]
            edited = True
            init_info.append("已创建模式文件。")
        if edited:
            wb.save(data_path)
        self.mode_upper_map = {k.upper(): k for k in self.mode_dict.keys()}
        return init_info

    async def can_process_msg(self, msg_str: str, meta: MessageMetaData) -> Tuple[bool, bool, Any]:
        # ── Task 3.3: 统一解析层前缀识别（仅做文本解析，不触库）──────────────
        if msg_str.startswith(".模式"):
            parse_result = _MODE_PARSER_ZH.parse(msg_str)
        elif msg_str.startswith(".mode"):
            parse_result = _MODE_PARSER_EN.parse(msg_str)
        else:
            return False, False, None

        # hint 携带 CommandParseResult 供 process_msg 消费（避免重复解析）
        # 群模式初始化检查移入 process_msg，保持 can_process_msg 纯解析、不触库
        return True, False, parse_result

    async def process_msg(self, msg_str: str, meta: MessageMetaData, hint: Any) -> List[BotCommandBase]:
        port = GroupMessagePort(
            meta.group_id) if meta.group_id else PrivateMessagePort(meta.user_id)
        # 判断功能开关
        if not self.bot.config.mode.enable:
            feedback = self.bot.loc_helper.format_loc_text(LOC_FUNC_DISABLE, func=self.readable_name)
            return [BotSendMsgCommand(self.bot.account, feedback, [port])]
        # 判断权限：群内需要权限>=0才能执行；私聊允许用户修改自己的私聊模式
        if meta.group_id and meta.permission < 0:
            feedback = self.bot.loc_helper.format_loc_text(LOC_PERMISSION_DENIED_NOTICE)
            return [BotSendMsgCommand(self.bot.account, feedback, [port])]

        # ── 构建本次 invocation 的唯一 CommandContext（per-invocation 缓存）──
        ctx = await CommandContextResolver.resolve(self.bot, meta)
        target_id = ctx.user_id if ctx.is_private else ctx.group_id
        is_private = ctx.is_private

        # ── 群模式初始化检查（原在 can_process_msg，移此处以符合"不触库"约定）──
        config = await ctx.group_config()
        current_mode = config.data.get("mode", "") if config else ""
        if current_mode == "":
            default_mode = self.bot.config.mode.default
            if default_mode != "":
                await self.switch_mode(target_id, default_mode, is_private=is_private)
            else:
                from core.data.models.extended import GroupConfig
                if config:
                    config.data["mode"] = "NULL"
                    await self.bot.db.group_config.upsert(config)
                else:
                    new_config = GroupConfig(group_id=ctx.config_key, data={"mode": "NULL"})
                    await self.bot.db.group_config.upsert(new_config)

        # ── Task 3.3: 从 CommandParseResult 消费解析结果 ───────────────────
        arg_var = hint.first_arg("").strip().upper()

        if arg_var == "DEFAULT" or arg_var == "CLEAR":
            feedback = await self.switch_mode(
                target_id, self.bot.config.mode.default, is_private=is_private)
        elif arg_var != "":
            feedback = await self.switch_mode(target_id, arg_var, is_private=is_private)
        else:
            # 显示当前目标（群/私聊）的模式（使用 ctx 缓存读取，避免重复 DB 访问）
            config = await ctx.group_config()

            stored_mode = config.data.get("mode", "") if config else ""

            # 处理空/NULL
            if not stored_mode or stored_mode == "NULL":
                stored_mode = self.bot.config.mode.default

            # 尝试从 mode_dict 中找到对应的显示信息
            mode_key = self.mode_upper_map.get(str(stored_mode).upper())

            if mode_key is not None:
                dice = self.mode_dict[mode_key][0] if len(self.mode_dict[mode_key]) > 0 else ""
                database = self.mode_dict[mode_key][1] if len(self.mode_dict[mode_key]) > 1 else ""
            else:
                # 回退到读取已保存的具体字段（若存在）或使用默认回退值
                dice = config.data.get("default_dice", "D20") if config else "D20"
                database = config.data.get("query_database", self.bot.config.query.private_database) if config else ""

            current_text = self.bot.loc_helper.format_loc_text(LOC_MODE_CURRENT, new_mode=stored_mode, dice=dice, database=database)
            list_text = self.bot.loc_helper.format_loc_text(LOC_MODE_LIST, modes="、".join(self.mode_dict.keys()))
            feedback = current_text + "\n" + list_text

        return [BotSendMsgCommand(self.bot.account, feedback, [port])]

    async def switch_mode(self, target_id: str, mode: str, is_private: bool = False) -> str:
        # 更新配置数据的内部异步函数
        # 私聊配置统一存储在 group_config 表，以 "__user__{user_id}" 作为 group_id
        async def update_group_config(tid: str, setting: List[str], var: List[str], is_private_inner: bool = False):
            config_key = f"__user__{tid}" if is_private_inner else tid
            config = await self.bot.db.group_config.get(config_key)
            if config is None:
                config = GroupConfig(group_id=config_key, data={})

            # 逐项更新字段
            for index in range(len(setting)):
                true_var: Any
                if isinstance(var[index], str) and var[index].isdigit():
                    true_var = int(var[index])
                elif isinstance(var[index], str) and var[index].upper() == "TRUE":
                    true_var = True
                elif isinstance(var[index], str) and var[index].upper() == "FALSE":
                    true_var = False
                else:
                    true_var = var[index]
                config.data[setting[index]] = true_var

            await self.bot.db.group_config.upsert(config)

        def guess_default_dice(db_name: str) -> str:
            """根据数据库名称猜测默认骰面"""
            db_upper = db_name.upper()
            # COC 系列 -> D100
            if re.search(r'COC', db_upper):
                return "100"
            # 忍神 / Shinobigami -> 2D6
            if '忍神' in db_name or re.search(r'SHINOBI', db_upper):
                return "2D6"
            # DND / PF / SF / SW 系列 -> D20
            if re.search(r'(DND|D&D|PF|PATHFINDER|SF|STARFINDER|SW|STARWARS)', db_upper):
                return "20"
            # 默认使用D20
            return "20"

        def find_matching_databases(query: str) -> List[str]:
            """在已连接数据库中查找匹配项"""
            query_upper = query.upper()
            results: List[str] = []
            for db_name in self.bot.db.query.list_databases():
                db_upper = db_name.upper()
                # 精确匹配（忽略大小写）
                if db_upper == query_upper:
                    return [db_name]  # 精确匹配直接返回
                # 模糊匹配：查询字符串是数据库名的子串
                if query_upper in db_upper:
                    results.append(db_name)
            return results

        matched = False
        feedback = ""
        # 尝试精准匹配预定义模式
        exact_key = self.mode_upper_map.get(mode)
        if exact_key is not None:
            await update_group_config(target_id, self.mode_field, [
                                exact_key]+self.mode_dict[exact_key], is_private_inner=is_private)
            feedback = self.bot.loc_helper.format_loc_text(
                LOC_MODE_SWITCH, new_mode=exact_key, dice=self.mode_dict[exact_key][0], database=self.mode_dict[exact_key][1])
            matched = True
        # 尝试模糊匹配预定义模式
        if not matched:
            result: List[str] = []
            for key in self.mode_dict.keys():
                ukey = key.upper()
                if mode in ukey:
                    result.append(ukey)
            if len(result) > 1:
                feedback = self.bot.loc_helper.format_loc_text(
                    LOC_MODE_LIKELY, modes="、".join(result))
                matched = True  # 有多个候选，不继续尝试数据库匹配
            elif len(result) == 1:
                key_upper = result[0]
                # 通过映射回原始 key 名称
                orig_key = self.mode_upper_map.get(key_upper)
                if orig_key is not None:
                    await update_group_config(target_id, self.mode_field, [
                                        orig_key]+self.mode_dict[orig_key], is_private_inner=is_private)
                    feedback = self.bot.loc_helper.format_loc_text(
                        LOC_MODE_SWITCH, new_mode=orig_key, dice=self.mode_dict[orig_key][0], database=self.mode_dict[orig_key][1])
                    matched = True

        # 如果预定义模式未匹配，尝试匹配已加载的数据库
        if not matched:
            db_matches = find_matching_databases(mode)
            if len(db_matches) == 1:
                # 唯一匹配，创建动态模式
                db_name = db_matches[0]
                dice = guess_default_dice(db_name)
                await update_group_config(target_id, self.mode_field, [
                                    mode, dice, db_name], is_private_inner=is_private)
                feedback = self.bot.loc_helper.format_loc_text(
                    LOC_MODE_DB_MATCH, database=db_name, dice=dice)
                matched = True
            elif len(db_matches) > 1:
                # 多个匹配，提示用户选择
                feedback = self.bot.loc_helper.format_loc_text(
                    LOC_MODE_DB_MULTI_MATCH, databases="、".join(db_matches))
                matched = True
            else:
                # 没有找到任何匹配
                feedback = self.bot.loc_helper.format_loc_text(
                    LOC_MODE_NOT_EXIST) + self.bot.loc_helper.format_loc_text(LOC_MODE_LIST, modes="、".join(self.mode_dict.keys()))

        return feedback

    def get_help(self, keyword: str, meta: MessageMetaData) -> str:
        if keyword == "config" or keyword == "mode":  # help后的接着的内容
            feedback: str = ".mode dnd/coc/ygo" \
                            "套用模式设置"
            return feedback
        return ""

    def get_description(self) -> str:
        return ".mode 模式系统"  # help指令中返回的内容