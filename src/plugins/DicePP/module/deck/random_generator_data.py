from typing import List, Tuple, Any, Dict, Optional, Set
from pathlib import Path
from enum import Enum
from datetime import datetime
import random
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment

from core.config import DATA_PATH
from module.roll import is_roll_exp, exec_roll_exp
from utils.time import get_current_date_raw, datetime_to_str_day
from utils.cq_code import get_cq_image
from utils.localdata import read_xlsx

RAND_SOURCE_FIELD_NAME = "生成器名称"
RAND_SOURCE_FIELD_VISIBLE = "是否可见"
RAND_SOURCE_FIELD_GLOBAL_PATH = "全局路径"
RAND_SOURCE_FIELD_RAND_TYPE = "随机方法"
RAND_SOURCE_FIELD_LIMIT_SINGLE = "单次上限"
RAND_SOURCE_FIELD_LIMIT_DAILY = "每日上限"
RAND_SOURCE_FIELD_FORMAT = "格式化"
RAND_SOURCE_FIELD_FORMAT_FUNC = "格式化函数"

RT_KEY_DAILY = "日期"
RT_KEY_USER = "个人"
RT_KEY_GROUP = "群组"
RT_KEY_LIST = [RT_KEY_DAILY, RT_KEY_USER, RT_KEY_GROUP]

RAND_SOURCE_FIELD_DEFAULT_COMMENT: Dict[str, Tuple[str, str]] = {
    RAND_SOURCE_FIELD_NAME: ("", "用户将通过类似.随机[生成器名称] 的方法使用该生成器, 如不可见则不会出现在生成器列表中展示给用户, 但依然可以通过指令使用."
                                 " 不填则为匿名生成器, 不可通过用户指令调用."),
    RAND_SOURCE_FIELD_VISIBLE: ("0", "1为对用户可见, 0为对用户隐藏(即只会通过其他随机生成器间接使用)"),
    RAND_SOURCE_FIELD_GLOBAL_PATH: ("", "不填则没有全局路径, 填入则代表使用该全局路径标识该生成器, 之后可在其他生成器中访问. 不能和其他全局路径重复."),
    RAND_SOURCE_FIELD_RAND_TYPE: ("", f"不填则使用默认随机方法, 填'{RT_KEY_DAILY}'代表每一天生成的随机内容都是一样的,"
                                      f" 填'{RT_KEY_USER}'代表根据个人账号决定结果,"
                                      f" 填'{RT_KEY_GROUP}'代表根据群聊号码决定结果,"
                                      " 填任意掷骰表达式则使用该掷骰表达式的结果从每个组别中根据从左往右的权重抽取结果. "
                                      " 可使用|组合, 如'日期|个人|d100'则可以起到类似jrrp的作用. 该项可以防止某些用户一直刷随机指令."
                                      "(这一项比较复杂, 还是不明白就到交流群问或者自己实验一下)"),
    RAND_SOURCE_FIELD_LIMIT_SINGLE: ("-1", "单次随机指令时的生成数据的次数上限, 超过则直接返回无效结果(包括被间接调用的次数) -1代表无限制"),
    RAND_SOURCE_FIELD_LIMIT_DAILY: ("-1", "一个用户每日最多使用的随机次数上限, 超过则直接返回无效结果(包括被间接调用的次数) -1代表无限制"),
    RAND_SOURCE_FIELD_FORMAT: ("", "格式化结果, 子条目必须被分到一个组别, 并以{组别}表示该组别生成的结果"),
    RAND_SOURCE_FIELD_FORMAT_FUNC: ("", "使用python函数格式化结果, 必须定义一个名为post_process的函数, 接受一个字符串参数输入, 输出结果字符串")

}
RAND_SOURCE_FIELD = list(RAND_SOURCE_FIELD_DEFAULT_COMMENT.keys())

RAND_ITEM_FIELD_NAME = "条目名称"
RAND_ITEM_FIELD_SOURCE = "数据来源"
RAND_ITEM_FIELD_WEIGHT = "组别与权重"
RAND_ITEM_FIELD_FLAG = "标志"
RAND_ITEM_FIELD_TRIGGER = "触发器"
RAND_ITEM_FIELD_CALLBACK = "后续生成器"
RAND_ITEM_FIELD_FORMAT = "格式化"

RAND_ITEM_FIELD_COMMENT = {
    RAND_ITEM_FIELD_NAME: "必填, 仅用来标注, 没有实际用途",
    RAND_ITEM_FIELD_SOURCE: "数据来源的路径, .或不填代表当前路径, 以/开头代表全局路径, 以文件夹结尾代表使用该生成器,"
                            " 以.xlsx结尾代表从该表格中的所有单元格作为数据来源, 以文件夹结尾则代表使用其中所有的.txt和图片作为数据来源",
    RAND_ITEM_FIELD_WEIGHT: "填入[组别][:权值], 如A, 或A:2, 同组的条目只会有一个被选中, 组别不填则为无组别, 权重不填或无效值则认为是1, 权重越大越有可能被选中",
    RAND_ITEM_FIELD_FLAG: "填入字符串标志, 可用&分隔多个标志, 当该条目被选中时将该标志设为开启, 应与组别以及触发器配合使用",
    RAND_ITEM_FIELD_TRIGGER: "填入字符串标志, 可用&分隔多个标志, 仅当填入的标志全部开启时可能被选中, 应与组别以及标志配合使用",
    RAND_ITEM_FIELD_CALLBACK: "填入其他生成器的全局路径, 如果该条目被触发, 则可以通过格式化结果将后续生成器的结果放入当前生成器的结果中 (请注意不要写出死循环)",
    RAND_ITEM_FIELD_FORMAT: "格式化结果, {_}或不填代表使用当前结果, {__}代表使用后续生成器得到的结果. 示例: 人类姓名:{_}",
}
RAND_ITEM_FIELD = list(RAND_ITEM_FIELD_COMMENT.keys())

ITEM_ROW_SPLIT_TOKEN = "*" * 12
IMG_SURFIX_LIST = [".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tif"]


class RandomSourceType(Enum):
    Literal = 0
    Directory = 1
    GlobalSource = 2
    Workbook = 3


class SourceFileType(Enum):
    TXT = 0
    IMG = 1


class RandomGenerateContext:
    """随机生成内容的上下文"""

    def __init__(self):
        self.user: str = ""
        self.group: str = ""
        self.time: datetime = get_current_date_raw()
        self.random_seed: str = ""
        self.flag_set: Set[str] = set()
        self.limit_dict: Dict[str, int] = {}
        self.daily_limit_dict: Dict[str, int] = {}


class RandomItem:
    """随机生成器的单个条目, 一次随机生成可由多个条目组合而成"""

    def __init__(self, name: str):
        self.name: str = name
        self.source: str = ""
        """
        可以为
        [GlobalSource]: 以/开头的其他生成器的GlobalPath
        [Directory]: ./或以./开头的文件夹(代表使用其中所有的.txt或图片)
        [Workbook]: 以./开头xlsx结尾的文件名(代表使用其中全部单元格作为数据来源)
        """
        self.group: str = ""
        self.weight: int = 1
        self.flag_set: Set[str] = set()
        self.trigger_set: Set[str] = set()
        self.format: str = ""
        self.next_gen_path: str = ""
        self.next_gen: Optional[RandomDataSource] = None
        self.index: int = -1
        self.root_path: Optional[Path] = None
        # 在resolve_source时赋值
        self.source_type: RandomSourceType = RandomSourceType.Literal
        self.auxiliary_data: Optional[Any] = None

    def resolve_source(self, global_source_dict: Dict[str, "RandomDataSource"]):
        """失败抛出 AssertionError"""
        if self.source.startswith("./"):
            # cases: [Directory] or [Workbook]
            # 尝试获取工作表名称
            relative_path = self.source[2:]
            sheet_name = ""
            if relative_path.rfind(".xlsx:") != -1:
                relative_path, sheet_name = relative_path.rsplit(":", maxsplit=1)
            # 获得最终路径
            resolved_path = self.root_path / relative_path
            assert resolved_path.exists(), f"路径{resolved_path}不存在"
            if resolved_path.is_dir():
                self.source_type = RandomSourceType.Directory
                self.auxiliary_data: List[List[Tuple[SourceFileType, Path]]] = load_source_info_from_path(resolved_path)
                assert self.auxiliary_data, f"资源目录{resolved_path}不含有可用文件"
            elif resolved_path.is_file() and resolved_path.suffix == ".xlsx":
                self.source_type = RandomSourceType.Workbook
                self.auxiliary_data: List[str] = []
                wb = read_xlsx(str(resolved_path.resolve()))
                if not sheet_name:
                    sheet_name = wb.sheetnames[0]
                assert sheet_name in wb.sheetnames, f"工作表{sheet_name}不存在"
                ws: Worksheet = wb[sheet_name]
                for col in ws.iter_cols(values_only=True):
                    self.auxiliary_data += [value.strip() for value in col if value and value.strip()]
                wb.close()
        elif self.source.startswith("/"):
            # [GlobalSource]
            assert self.source in global_source_dict, f"全局路径{self.source}不存在"
            self.source_type = RandomSourceType.GlobalSource
            self.auxiliary_data = global_source_dict[self.source]
        elif not self.source:
            self.source_type = RandomSourceType.Literal
            self.auxiliary_data = self.name
        else:
            assert False, f"无效的资源类型"

        if self.next_gen_path:
            assert self.next_gen_path in global_source_dict
            self.next_gen = global_source_dict[self.next_gen_path]

    def gen_result(self, context: RandomGenerateContext) -> str:
        result: str
        assert self.auxiliary_data

        if self.source_type == RandomSourceType.GlobalSource:
            target_source: RandomDataSource = self.auxiliary_data
            result = target_source.gen_result(context)
        elif self.source_type == RandomSourceType.Directory:
            file_info_list: List[List[Tuple[SourceFileType, Path]]] = self.auxiliary_data
            result_info: List[Tuple[SourceFileType, Path]] = random.choice(file_info_list)
            result = ""
            for file_type, file_path in result_info:
                if not file_path.exists():
                    return f"数据文件{file_path.relative_to(DATA_PATH)}丢失"
                if file_type == SourceFileType.TXT:
                    result += file_path.read_text()
                elif file_type == SourceFileType.IMG:
                    result += get_cq_image(file_path.read_bytes())
                else:
                    return "无效文件类型"
        elif self.source_type == RandomSourceType.Workbook:
            text_list: List[str] = self.auxiliary_data
            result = random.choice(text_list)
        elif self.source_type == RandomSourceType.Literal:
            result: str = self.auxiliary_data
        else:
            return "无效资源类型"

        next_result: str = ""
        if self.next_gen and self.format:
            next_result = self.next_gen.gen_result(context)
        if self.format:
            result = self.format.format(_=result, __=next_result)
        return result


class RandomDataSource:
    """随机生成器的数据来源"""

    def __init__(self, name: str, path: Path):
        self.name: str = name
        self.root_path: Path = path
        self.is_visible: bool = False
        self.global_path: str = ""
        self.random_seed_flag: int = 0  # Bits: 1: Daily 2: UserId 3: GroupId
        self.random_seed_exp: str = ""
        self.limit_single: int = -1
        self.limit_daily: int = -1
        self.format: str = ""
        self.format_func: str = ""
        self.items: List[RandomItem] = []
        self.items_in_group: Dict[str, List[RandomItem]] = {}

    def write_to_sheet(self, target: Worksheet):
        """不会抛出异常"""
        # 先清空表格
        target.delete_rows(1, target.max_row)
        # 写入根规则
        for row_index, (field, info) in enumerate(RAND_SOURCE_FIELD_DEFAULT_COMMENT.items()):
            row_index += 1
            val_text, comment = info
            # 写入参数名称
            cell_field = target.cell(row=row_index, column=1, value=field)
            cell_field.comment = Comment(comment, "DicePP")
            # 生成参数
            if field == RAND_SOURCE_FIELD_NAME:
                val_text = self.name
            elif field == RAND_SOURCE_FIELD_VISIBLE:
                val_text = "1" if self.is_visible else "0"
            elif field == RAND_SOURCE_FIELD_GLOBAL_PATH:
                val_text = self.global_path
            elif field == RAND_SOURCE_FIELD_RAND_TYPE:
                val_text = ""
                if self.random_seed_flag:
                    for bit_index, flag_name in enumerate(RT_KEY_LIST):
                        if self.random_seed_flag & (1 << (bit_index + 1)):
                            val_text += flag_name + "|"
                    val_text = val_text[:-1]
                if self.random_seed_exp:
                    val_text = f"{val_text}|{self.random_seed_exp}" if val_text else self.random_seed_exp
            elif field == RAND_SOURCE_FIELD_LIMIT_SINGLE:
                val_text = str(self.limit_single)
            elif field == RAND_SOURCE_FIELD_LIMIT_DAILY:
                val_text = str(self.limit_daily)
            elif field == RAND_SOURCE_FIELD_FORMAT:
                val_text = self.format
            elif field == RAND_SOURCE_FIELD_FORMAT_FUNC:
                val_text = self.format_func
            # 写入参数
            target.cell(row=row_index, column=2, value=val_text)
        # 写入分隔符
        first_item_row = len(RAND_SOURCE_FIELD_DEFAULT_COMMENT) + 3
        target.cell(row=first_item_row-1, column=1, value=ITEM_ROW_SPLIT_TOKEN)
        # 写入条目
        for row_index, (field, info) in enumerate(RAND_ITEM_FIELD_COMMENT.items()):
            row_index = first_item_row + row_index
            comment = info
            # 写入参数名称
            cell_field = target.cell(row=row_index, column=1, value=field)
            cell_field.comment = Comment(comment, "DicePP")
        for column_index, rand_item in enumerate(self.items):
            column_index += 2
            for row_index, (field, info) in enumerate(RAND_ITEM_FIELD_COMMENT.items()):
                row_index = first_item_row + row_index
                val_text = ""
                # 生成参数值
                if field == RAND_ITEM_FIELD_NAME:
                    val_text = rand_item.name
                elif field == RAND_ITEM_FIELD_SOURCE:
                    val_text = rand_item.source
                elif field == RAND_ITEM_FIELD_WEIGHT:
                    if rand_item.group:
                        val_text = rand_item.group
                        if rand_item.weight != 1:
                            val_text = f"{val_text}:{rand_item.weight}"
                elif field == RAND_ITEM_FIELD_FLAG:
                    if rand_item.flag_set:
                        val_text = "&".join(list(rand_item.flag_set))
                elif field == RAND_ITEM_FIELD_TRIGGER:
                    if rand_item.trigger_set:
                        val_text = "&".join(list(rand_item.trigger_set))
                elif field == RAND_ITEM_FIELD_CALLBACK:
                    val_text = rand_item.next_gen_path
                elif field == RAND_ITEM_FIELD_FORMAT:
                    val_text = rand_item.format
                # 写入参数
                target.cell(row=row_index, column=column_index, value=val_text)

    def read_from_sheet(self, target: Worksheet) -> str:
        """不会抛出异常, 返回错误信息, 返回空字符串说明读取成功"""
        # 读取根规则
        first_item_row = -1
        for row in target.iter_rows(min_row=1):
            row: Tuple[Cell]
            title = row[0].value
            if title in RAND_SOURCE_FIELD_DEFAULT_COMMENT:
                try:
                    assert row[1].value
                    val_text = str(row[1].value)
                    if title not in [RAND_SOURCE_FIELD_FORMAT, RAND_SOURCE_FIELD_FORMAT_FUNC]:
                        val_text = val_text.strip()
                except (IndexError, AssertionError):
                    val_text = ""
                if title == RAND_SOURCE_FIELD_NAME:
                    self.name = val_text
                elif title == RAND_SOURCE_FIELD_VISIBLE:
                    self.is_visible = (val_text == "1")
                elif title == RAND_SOURCE_FIELD_GLOBAL_PATH:
                    self.global_path = val_text
                elif title == RAND_SOURCE_FIELD_RAND_TYPE:
                    flag_list = [flag.strip() for flag in val_text.split("|")]
                    for flag in flag_list:
                        if not flag:
                            continue
                        if flag in RT_KEY_LIST:
                            self.random_seed_flag += 1 << (RT_KEY_LIST.index(flag) + 1)
                        else:  # 掷骰表达式
                            if is_roll_exp(flag):
                                self.random_seed_exp = flag
                            else:
                                return f"随机方法{flag}是不合法的掷骰表达式"
                elif title == RAND_SOURCE_FIELD_LIMIT_SINGLE:
                    try:
                        self.limit_single = int(val_text)
                    except ValueError:
                        self.limit_single = -1
                elif title == RAND_SOURCE_FIELD_LIMIT_DAILY:
                    try:
                        self.limit_daily = int(val_text)
                    except ValueError:
                        self.limit_daily = -1
                elif title == RAND_SOURCE_FIELD_FORMAT:
                    self.format = val_text
                elif title == RAND_SOURCE_FIELD_FORMAT_FUNC:
                    if val_text.startswith("def post_process"):
                        try:
                            exec_global = {"result": "test_str"}
                            exec(val_text + "\nresult = post_process(result)", exec_global)
                            assert type(exec_global["result"]) is str
                            self.format_func = val_text
                        except Exception as e:
                            return f"自定义格式化函数不正确: {e}"

            elif title == ITEM_ROW_SPLIT_TOKEN:
                first_item_row = row[0].row + 1
                break
            elif title in RAND_ITEM_FIELD_COMMENT:
                first_item_row = row[0].row
                break
        if first_item_row == -1:
            return f"该工作表格式不正确"
        # 读取条目
        self.items = []
        for row in target.iter_rows(min_row=first_item_row):
            row: Tuple[Cell]
            title = row[0].value
            if title == RAND_ITEM_FIELD_NAME:
                # 初始化条目
                found_empty_cell = False
                for cell in row[1:]:
                    cell: Cell
                    if cell.value:
                        new_item = RandomItem(str(cell.value))
                        new_item.index = len(self.items)
                        new_item.root_path = self.root_path
                        self.items.append(new_item)
                        if found_empty_cell:
                            return "所有条目必须首先填入名称且连续, 中间不能存在空白条目"
                    else:
                        found_empty_cell = True
                continue
            # 其他属性
            if not self.items:
                return f"{RAND_ITEM_FIELD_NAME}必须作为首个条目属性"
            val_text_list = [(str(cell.value) if cell.value else "") for cell in row[1:len(self.items)+1]]
            if title != RAND_ITEM_FIELD_FORMAT:
                val_text_list = [val_text.strip() for val_text in val_text_list]
            assert len(val_text_list) == len(self.items)
            if title not in RAND_ITEM_FIELD_COMMENT:
                return f"{title}为无效的条目属性"
            for item_index, rand_item in enumerate(self.items):
                val_text = val_text_list[item_index]
                if title == RAND_ITEM_FIELD_SOURCE:
                    if val_text:
                        rand_item.source = val_text
                elif title == RAND_ITEM_FIELD_WEIGHT:
                    if val_text:
                        if ":" in val_text:
                            rand_item.group, weight_str = val_text.split(":", maxsplit=1)
                            try:
                                rand_item.weight = int(weight_str)
                                assert rand_item.weight > 0
                            except (ValueError, AssertionError):
                                return f"{rand_item.name}的权重({weight_str})必须为正整数"
                        else:
                            rand_item.group = val_text
                elif title == RAND_ITEM_FIELD_FLAG:
                    if val_text:
                        rand_item.flag_set = set(val_text.split("&"))
                elif title == RAND_ITEM_FIELD_TRIGGER:
                    if val_text:
                        rand_item.trigger_set = set(val_text.split("&"))
                elif title == RAND_ITEM_FIELD_CALLBACK:
                    rand_item.next_gen_path = val_text
                elif title == RAND_ITEM_FIELD_FORMAT:
                    rand_item.format = val_text

        # 将item按照group分类
        self.items_in_group = {}
        for item in self.items:
            if item.group not in self.items_in_group:
                self.items_in_group[item.group] = [item]
            else:
                self.items_in_group[item.group].append(item)
        if self.name:
            target.title = self.name
        # 检查格式化是否正确
        if self.format:
            try:
                dumb_info = {name: "ABC" for name in self.items_in_group.keys()}
                self.format.format(**dumb_info)
            except KeyError:
                return f"随机生成器{self.name}格式化规则不正确: {self.format} 可用群组:{list(self.items_in_group.keys())}"
        return ""

    def resolve_source(self, global_source_dict: Dict[str, "RandomDataSource"]):
        """失败抛出 AssertionError """
        for item in self.items:
            try:
                item.resolve_source(global_source_dict)
            except AssertionError as e:
                raise AssertionError(f"{self.name}:{item.name}:{e.args}")

    def gen_result(self, context: RandomGenerateContext) -> str:
        # 记录上层随机数种子与(可能)生成新的随机数种子
        prev_seed = context.random_seed
        if not context.random_seed and self.random_seed_flag != 0:
            context.random_seed = gen_new_seed(self.random_seed_flag, context)
            random.seed(context.random_seed)

        # 找到需要生成的条目
        used_item: List[RandomItem] = []
        for group, item_list in self.items_in_group.items():
            filtered_item_list = []
            for item in item_list:
                if not item.trigger_set or item.trigger_set.intersection(context.flag_set):
                    filtered_item_list.append(item)
            if filtered_item_list:
                if group:  # 从组内随机选一个条目
                    filtered_item_list = [select_item_from_group(filtered_item_list, self.random_seed_exp)]
                used_item += filtered_item_list
                for item in filtered_item_list:
                    if item.flag_set:
                        context.flag_set.update(item.flag_set)
        # 统计限制次数
        if self.limit_single > 0:
            limit_key = self.name + self.global_path
            if context.limit_dict.get(limit_key, 0) >= self.limit_single:
                return "[超出单次限制]"
            context.limit_dict[limit_key] = context.limit_dict.get(limit_key, 0) + 1
        if self.limit_daily > 0:
            limit_key = self.name + self.global_path
            if context.daily_limit_dict.get(limit_key, 0) >= self.limit_daily:
                return "[超出每日限制]"
            context.daily_limit_dict[limit_key] = context.daily_limit_dict.get(limit_key, 0) + 1

        # 生成结果
        final_result_list: List[Tuple[RandomItem, str]] = [(item, item.gen_result(context)) for item in used_item]

        # 格式化结果
        if self.format:
            result_in_group: Dict[str, str] = {item.group: text for item, text in final_result_list}
            try:
                result = self.format.format(**result_in_group)
            except KeyError:
                result = f"随机生成器{self.name}格式化规则不正确: {self.format} {result_in_group}"
        else:
            if len(final_result_list) > 1:
                final_result_list = sorted(final_result_list, key=lambda x: x[0].index)
                result = "".join([t[1] for t in final_result_list])
            elif len(final_result_list) == 1:
                result = final_result_list[0][1]
            else:
                result = ""
        # 使用自定义格式化函数
        if self.format_func:
            exec_global = {"result": result}
            exec(self.format_func + "\nresult = post_process(result)", exec_global)
            result = exec_global["result"]

        # 恢复上层随机数种子, 如果之前就有随机数种子, 留给上层处理
        if not prev_seed:
            random.seed(random.random())
            context.random_seed = prev_seed
        return result


def gen_new_seed(random_seed_flag: int, context: RandomGenerateContext) -> str:
    new_seed = ""
    if flag_include(random_seed_flag, RT_KEY_DAILY):
        new_seed += datetime_to_str_day(context.time)
    if flag_include(random_seed_flag, RT_KEY_USER):
        new_seed += context.user
    if flag_include(random_seed_flag, RT_KEY_GROUP):
        new_seed += context.group
    return new_seed


def flag_include(random_seed_flag: int, flag_name: str) -> bool:
    assert flag_name in RT_KEY_LIST
    return (random_seed_flag & (1 << (RT_KEY_LIST.index(flag_name) + 1))) != 0


def select_item_from_group(item_list: List[RandomItem], weight_exp: str = "") -> RandomItem:
    assert len(item_list) > 0
    cur_weight: int
    if weight_exp:
        cur_weight = exec_roll_exp(weight_exp).get_val()
    else:
        # 根据组内权重总和计算
        weight_max = sum([item.weight for item in item_list])
        cur_weight = random.randint(1, weight_max)
    for item in item_list:
        if cur_weight <= item.weight:
            return item
        cur_weight -= item.weight
    return item_list[-1]


def load_source_info_from_path(resolved_path: Path) -> List[List[Tuple[SourceFileType, Path]]]:
    all_source_file: List[Tuple[SourceFileType, Path]] = []
    for file_path in resolved_path.iterdir():
        if not file_path.is_file():
            continue
        file_surfix = file_path.suffix.lower()
        if file_surfix == ".txt":
            all_source_file.append((SourceFileType.TXT, file_path))
        elif file_surfix in IMG_SURFIX_LIST:
            all_source_file.append((SourceFileType.IMG, file_path))
    # 通过前缀进行分组
    all_source_in_group: Dict[str, List[Tuple[int, Tuple[SourceFileType, Path]]]] = {}
    for source_file in all_source_file:
        group_name = source_file[1].stem
        in_group_index = ""
        if "." in group_name:
            source_file_name, in_group_index = group_name.rsplit(".", maxsplit=1)
        group_name += source_file[1].suffix
        if group_name not in all_source_in_group:
            all_source_in_group[group_name] = []
        try:
            in_group_index = int(in_group_index)
        except ValueError:
            in_group_index = 0
        all_source_in_group[group_name].append((in_group_index, source_file))
    final_source_info: List[List[Tuple[SourceFileType, Path]]] = []
    for in_group_list in all_source_in_group.values():
        if len(in_group_list) == 1:
            final_source_info.append([in_group_list[0][1]])
        else:
            sorted_group_list = [source_info for index, source_info in sorted(in_group_list, key=lambda x: x[0])]
            final_source_info.append(sorted_group_list)
    return final_source_info
