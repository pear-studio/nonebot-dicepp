from typing import Dict, Optional
import os
import re
import random

import openpyxl
from openpyxl.comments import Comment
from openpyxl.worksheet import worksheet

from utils.logger import dice_log
from utils.localdata import read_xlsx
from core.config import DATA_PATH as ROOT_DATA_PATH
from core.communication import preprocess_msg

from core.localization.common import COMMON_LOCAL_TEXT, COMMON_LOCAL_COMMENT
from core.localization.localization_text import LocalizationText

LOCAL_FILE_PATH = "localization.xlsx"
CHAT_FILE_PATH = "chat.xlsx"

DEFAULT_ID = "Default"

DEFAULT_CHAT_KEY = "^hi$"
DEFAULT_CHAT_TEXT = ["Hello", "G'Day"]
DEFAULT_CHAT_COMMENT = "可以使用正则表达式匹配, 大小写不敏感; 后面接着想要的回复, 有多个回复则会随机选择一个"


class LocalizationManager:
    def __init__(self, data_path: str, identifier: str):
        self.data_path = os.path.join(data_path, LOCAL_FILE_PATH)
        self.chat_data_path = os.path.join(data_path, CHAT_FILE_PATH)
        self.identifier = identifier
        self.all_local_texts: Dict[str, LocalizationText] = {}
        self.all_chat_texts: Dict[str, LocalizationText] = {}

        # 通用的本地化语句
        for key in COMMON_LOCAL_TEXT.keys():
            self.register_loc_text(key, COMMON_LOCAL_TEXT[key], COMMON_LOCAL_COMMENT[key])

    def load_localization(self):
        """用xlsx里的配置覆写之前的本地化配置, 不存在本地文件时直接返回"""
        workbook, local_sheet = load_sheet_from_path(self.data_path, self.identifier)
        if not workbook:
            dice_log(f"[Local] [Load] 无法找到本地化文件 {self.data_path.replace(ROOT_DATA_PATH, '~')} {self.identifier}")
            return
        for row in local_sheet.iter_rows():
            key = str(row[0].value)  # 第一个元素为关键字
            if key not in self.all_local_texts:  # 无效的关键字
                continue
            comment: str = self.all_local_texts[key].comment  # 沿用原来的注释, 不用文件里的
            self.all_local_texts[key] = LocalizationText(key, comment=comment)
            for text in [str(cell.value) for cell in row[1:] if cell.value and cell.value.strip()]:
                self.all_local_texts[key].add(text)
        dice_log(f"[Local] [Load] 成功读取本地化文件 {self.data_path.replace(ROOT_DATA_PATH, '~')}")
        workbook.close()

    def save_localization(self):
        """注意按多个机器人会读写同一个配置文件, 如果并行可能存在写冲突, 现在单线程异步没问题"""
        workbook, local_sheet = get_sheet_from_path(self.data_path, self.identifier)
        for ri, loc_text in enumerate(self.all_local_texts.values()):
            save_loc_text_to_row(local_sheet, loc_text, ri + 1)

        try:
            workbook.save(self.data_path)
            dice_log(f"[Local] [Save] 成功更新本地化文件 {self.data_path.replace(ROOT_DATA_PATH, '~')} {local_sheet.title}")
        except PermissionError:
            dice_log(f"[Local] [Save] 无法保存本地化文件 {self.data_path.replace(ROOT_DATA_PATH, '~')}, 没有写入权限")
        workbook.close()

    def load_chat(self):
        """从xlsx中读取自定义对话文件"""
        def add_default_chat():
            """增加默认自定义对话"""
            self.all_chat_texts[DEFAULT_CHAT_KEY] = LocalizationText(DEFAULT_CHAT_KEY, comment=DEFAULT_CHAT_COMMENT)
            for default_text in DEFAULT_CHAT_TEXT:
                self.all_chat_texts[DEFAULT_CHAT_KEY].add(default_text)

        workbook, chat_sheet = load_sheet_from_path(self.chat_data_path, self.identifier)
        if not workbook:
            dice_log(f"[Local] [ChatLoad] 无法找到自定义对话文件 {self.chat_data_path.replace(ROOT_DATA_PATH, '.')} {self.identifier}")
            add_default_chat()
            return

        for row in chat_sheet.iter_rows():
            key = str(row[0].value)  # 第一个元素为关键字
            key = preprocess_msg(key)  # 对key做一下预处理, 因为匹配的目标是预处理过后的
            comment: str = row[0].comment  # 沿用文件里的注释
            self.all_chat_texts[key] = LocalizationText(key, comment=comment)
            for text in [str(cell.value) for cell in row[1:] if cell.value and cell.value.strip()]:
                self.all_chat_texts[key].add(text)

        has_chat: bool = (len(self.all_chat_texts) != 0)
        if not has_chat:
            add_default_chat()
        dice_log(f"[Local] [ChatLoad] 成功读取本地化文件 {self.chat_data_path.replace(ROOT_DATA_PATH, '~')}")
        workbook.close()

    def save_chat(self):
        """注意多个机器人会读写同一个配置文件, 如果并行可能存在写冲突, 现在单线程异步没问题"""
        workbook, local_sheet = get_sheet_from_path(self.chat_data_path, self.identifier)
        for ri, loc_text in enumerate(self.all_chat_texts.values()):
            save_loc_text_to_row(local_sheet, loc_text, ri + 1)

        try:
            workbook.save(self.chat_data_path)
            dice_log(f"[Local] [ChatSave] 成功更新自定义对话文件 {self.chat_data_path.replace(ROOT_DATA_PATH, '~')} {local_sheet.title}")
        except PermissionError:
            dice_log(f"[Local] [ChatSave] 无法保存自定义对话文件 {self.chat_data_path.replace(ROOT_DATA_PATH, '~')}, 没有写入权限")
        workbook.close()

    def register_loc_text(self, key: str, default_text: str, comment: str = ""):
        """
        将一个本地化语句注册至Helper中
        Args:
            key: 本地化语句的关键字
            default_text: 本地化语句的默认值
            comment: 对本地化语句的注释
        """
        self.all_local_texts[key] = LocalizationText(key, default_text, comment)

    def get_loc_text(self, key: str) -> LocalizationText:
        """
        获取本地化语句
        Args:
            key: 本地化语句的关键字
        """
        return self.all_local_texts[key]

    def format_loc_text(self, key: str, **kwargs) -> str:
        """
        格式化并返回本地化语句
        Args:
            key: 本地化语句的关键字
            **kwargs: 本地化语句需要的参数, 可以传不会用到的参数
        """
        loc_text: LocalizationText = self.get_loc_text(key)
        if kwargs:
            return loc_text.get().format(**kwargs)
        else:
            return loc_text.get()

    def process_chat(self, msg: str, **kwargs) -> str:
        """
        Args:
            msg: 用户输入的语句
            **kwargs: 目前用不到

        Returns:
            如果msg能与任意自定义聊天关键字匹配, 返回一个随机回复, 否则返回空字符串
        """
        valid_loc_text_list = []
        for key, loc_text in self.all_chat_texts.items():
            result = re.match(key, msg)
            if result:
                valid_loc_text_list.append(loc_text)

        loc_text: Optional[LocalizationText] = random.choice(valid_loc_text_list) if valid_loc_text_list else None
        if loc_text:
            if kwargs:
                return loc_text.get().format(**kwargs)
            else:
                return loc_text.get()
        return ""


def save_loc_text_to_row(sheet: worksheet, l_text: LocalizationText, row: int):
    # 先清空旧数据
    sheet.delete_rows(idx=row)
    sheet.insert_rows(idx=row)
    # 加入新数据
    header = sheet.cell(row=row, column=1, value=l_text.key)
    if l_text.comment:
        header.comment = Comment(l_text.comment, "DicePP")

    for ci, text in enumerate(l_text.loc_texts):
        sheet.cell(row=row, column=ci + 2, value=text)


def load_sheet_from_path(data_path: str, identifier: str, default_id: str = DEFAULT_ID) -> (openpyxl.Workbook, worksheet):
    """若指定data_path无效或id无效, 返回None. 若id无效会尝试使用default_id, 一般用来得到读取的sheet"""
    if not os.path.exists(data_path):
        return None, None
    workbook = read_xlsx(data_path)
    if identifier in workbook.sheetnames:
        sheet = workbook[identifier]
        sheet.title = identifier
    elif default_id in workbook.sheetnames:
        sheet = workbook[default_id]
        sheet.title = default_id
    else:
        workbook.close()
        return None, None
    return workbook, sheet


def get_sheet_from_path(data_path: str, identifier: str) -> (openpyxl.Workbook, worksheet):
    """若指定的data_path无效或id无效, 就创建新的workbook或worksheet. 一般用来得到写入的sheet"""
    feedback: str
    if os.path.exists(data_path):
        workbook = read_xlsx(data_path)
    else:
        workbook = openpyxl.Workbook()
        for name in workbook.sheetnames:
            del workbook[name]

    if identifier in workbook.sheetnames:
        sheet = workbook[identifier]
    else:
        sheet = workbook.create_sheet(identifier)
    sheet.title = identifier
    return workbook, sheet
