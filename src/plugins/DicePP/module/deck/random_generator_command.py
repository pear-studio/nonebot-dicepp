"""
随机生成器指令, 从资料库中随机生成材料并回复给用户
"""
import os
from typing import List, Tuple, Any, Dict, Optional
from pathlib import Path
import openpyxl

from core.bot import Bot
from core.config import DATA_PATH
from core.command.const import *
from core.command import UserCommandBase, custom_user_command
from core.command import BotCommandBase, BotSendMsgCommand
from core.communication import MessageMetaData, PrivateMessagePort, GroupMessagePort
from utils.localdata import read_xlsx, update_xlsx
from utils.string import match_substring

from module.deck.random_generator_data import RandomDataSource, RandomGenerateContext

LOC_RAND_GEN_LIST = "rand_gen_list"
LOC_RAND_GEN_MISS = "rand_gen_miss"
LOC_RAND_GEN_VAGUE = "rand_gen_vague"

CFG_RAND_GEN_ENABLE = "random_gen_enable"
CFG_RAND_GEN_DATA_PATH = "random_gen_data_path"
RAND_GEN_DATA_PATH = "RandomGenData"
META_FILE_NAME = "rule.xlsx"


@custom_user_command(readable_name="随机生成器指令", priority=DPP_COMMAND_PRIORITY_DEFAULT,
                     flag=DPP_COMMAND_FLAG_DRAW)
class RandomGeneratorCommand(UserCommandBase):

    def __init__(self, bot: Bot):
        super().__init__(bot)
        self.source_list: List[RandomDataSource] = []
        self.source_name_dict: Dict[str, RandomDataSource] = {}
        self.daily_limit_dict: Dict[str, Dict[str, int]] = {}

        bot.loc_helper.register_loc_text(LOC_RAND_GEN_LIST, "These are available generator: {list}", "展示可用的随机生成器列表")
        bot.loc_helper.register_loc_text(LOC_RAND_GEN_MISS, "Cannot find {name} generator...", "找不到用户输入的随机生成器")
        bot.loc_helper.register_loc_text(LOC_RAND_GEN_VAGUE, "Maybe you want these generator? {list}", "用户输入的随机生成器名称有多个匹配可能")

        bot.cfg_helper.register_config(CFG_RAND_GEN_ENABLE, "1", "随机生成器指令开关")
        bot.cfg_helper.register_config(CFG_RAND_GEN_DATA_PATH, f"./{RAND_GEN_DATA_PATH}", "随机生成器指令的数据来源, .代表Data文件夹")

    def delay_init(self) -> List[str]:
        # 从本地文件中读取资料
        data_path_list: List[str] = self.bot.cfg_helper.get_config(CFG_RAND_GEN_DATA_PATH)
        for i, path in enumerate(data_path_list):
            if path.startswith("./"):  # 用DATA_PATH作为当前路径
                data_path_list[i] = os.path.join(DATA_PATH, path[2:])
            data_path_list[i] = Path(data_path_list[i])
        data_path_list: List[Path]
        data_dir_path_list: List[Path] = []
        for data_path in data_path_list:
            if data_path.exists():
                data_dir_path_list += [path for path in data_path.iterdir() if path.is_dir()]
            else:
                data_path.mkdir(parents=True)
        init_info: List[str] = []
        self.init_from_data_dir(data_dir_path_list, init_info)
        self.finalize_init(init_info)
        # init_info.append(self.get_state())
        return init_info

    def can_process_msg(self, msg_str: str, meta: MessageMetaData) -> Tuple[bool, bool, Any]:
        should_proc: bool = msg_str.startswith(".随机")
        should_pass: bool = False
        return should_proc, should_pass, msg_str[3:].strip()

    def process_msg(self, msg_str: str, meta: MessageMetaData, hint: Any) -> List[BotCommandBase]:
        port = GroupMessagePort(meta.group_id) if meta.group_id else PrivateMessagePort(meta.user_id)
        # 判断功能开关
        try:
            assert (int(self.bot.cfg_helper.get_config(CFG_RAND_GEN_ENABLE)[0]) != 0)
        except AssertionError:
            feedback = self.bot.loc_helper.format_loc_text(CFG_RAND_GEN_ENABLE, func=self.readable_name)
            return [BotSendMsgCommand(self.bot.account, feedback, [port])]
        # 解析语句
        arg_str: str = hint
        feedback: str = ""
        if not arg_str:  # 展示列表
            gen_list: List[str] = []
            for name, source in self.source_name_dict.items():
                if source.is_visible:
                    gen_list.append(name)
            feedback = self.format_loc(LOC_RAND_GEN_LIST, list=", ".join(gen_list))
        elif arg_str == "列表":  # 展示完整列表
            gen_list: List[Tuple[str, str]] = []
            for name, source in self.source_name_dict.items():
                gen_list.append((name, source.global_path))
            result = ", ".join([info[0] for info in sorted(gen_list, key=lambda x: x[1])])
            feedback = self.format_loc(LOC_RAND_GEN_LIST, list=result)
        else:
            time = 1
            if "#" in arg_str:
                time_str, arg_str = arg_str.split("#", maxsplit=1)
                try:
                    time = int(time_str)
                    assert 0 < time <= 10
                except (ValueError, AssertionError):
                    time = 1

            target_source_name = arg_str
            target_source: Optional[RandomDataSource] = None
            poss_source_names = match_substring(target_source_name, self.source_name_dict.keys())
            if target_source_name in poss_source_names:
                poss_source_names = [target_source_name]
            if len(poss_source_names) == 0:
                feedback += self.format_loc(LOC_RAND_GEN_MISS, name=target_source_name) + "\n"
            elif len(poss_source_names) > 1:
                feedback += self.format_loc(LOC_RAND_GEN_VAGUE, list=", ".join(poss_source_names)) + "\n"
            else:
                target_source_name = poss_source_names[0]
                target_source = self.source_name_dict[target_source_name]

            if target_source:
                context = RandomGenerateContext()
                context.user = meta.user_id
                context.group = meta.group_id
                context.daily_limit_dict = self.daily_limit_dict.get(meta.user_id, {})
                if time == 1:
                    feedback = target_source.gen_result(context)
                else:
                    feedback = ""
                    for t in range(time):
                        feedback += f"#{t+1} {target_source.gen_result(context)}\n"
                        context.flag_set = set()
                    feedback = feedback.strip()
                self.daily_limit_dict[meta.user_id] = context.daily_limit_dict

        return [BotSendMsgCommand(self.bot.account, feedback, [port])]

    def tick_daily(self) -> List[BotCommandBase]:
        self.daily_limit_dict = {}
        return []

    def get_help(self, keyword: str, meta: MessageMetaData) -> str:
        if keyword == "随机":  # help后的接着的内容
            feedback: str = ".随机 查看常用随机生成器列表\n.随机列表 查看所有生成器列表\n" \
                            ".随机[生成器名称] 使用随机生成器"
            return feedback
        return ""

    def get_description(self) -> str:
        return ".随机 随机生成各种数据"  # help指令中返回的内容

    def init_from_data_dir(self, data_dir_path_list: List[Path], error_info: List[str]):
        sub_data_dir_path_list: List[Path] = []
        for dir_path in data_dir_path_list:
            has_meta_file = False
            for path in dir_path.iterdir():
                if path.is_dir():
                    sub_data_dir_path_list.append(path)
                elif path.name == META_FILE_NAME:
                    has_meta_file = True
                    self.process_meta_file(path, error_info)
            if not has_meta_file:
                self.create_meta_file(dir_path)
        if sub_data_dir_path_list:
            self.init_from_data_dir(sub_data_dir_path_list, error_info)

    def create_meta_file(self, dir_path: Path):
        wb = openpyxl.Workbook()
        for name in wb.sheetnames:
            del wb[name]
        ws_temp = wb.create_sheet("生成器规则")
        new_source = RandomDataSource(dir_path.name, dir_path)
        new_source.write_to_sheet(ws_temp)
        self.source_list.append(new_source)
        update_xlsx(wb, str((dir_path / META_FILE_NAME).resolve()))
        wb.close()
        return wb

    def process_meta_file(self, meta_path: Path, error_info: List[str]):
        assert meta_path.suffix == ".xlsx"
        try:
            wb = read_xlsx(str(meta_path.resolve()))
        except PermissionError:
            error_info.append(f"读取{meta_path}时遇到错误: 权限不足")
            return
        for name in wb.sheetnames:
            ws = wb[name]
            new_source = RandomDataSource("", meta_path.parent)
            error = new_source.read_from_sheet(ws)
            if error:
                error = f"读取{meta_path.relative_to(DATA_PATH)}/{name}时遇到错误: {error}"
                error_info.append(error)
            else:
                self.source_list.append(new_source)
                new_source.write_to_sheet(ws)
        update_xlsx(wb, str(meta_path.resolve()))
        wb.close()

    def finalize_init(self, error_info: List[str]):
        """分析各个source之间的引用关系, 预读取需要的xlsx中的数据, 分析txt和图片等文件的路径"""
        global_source_dict: Dict[str, RandomDataSource] = {}
        for source in self.source_list:
            if source.global_path:
                global_source_dict[source.global_path] = source
        invalid_source = []
        first_time = True
        while invalid_source or first_time:
            invalid_source, first_time = [], False
            for source in self.source_list:
                try:
                    source.resolve_source(global_source_dict)
                except AssertionError as e:
                    error_info.append(f"生成器初始化失败:{source.name}{e.args[0]}")
                    invalid_source.append(source)
            for source in invalid_source:
                self.source_list.remove(source)
                if source.global_path in global_source_dict:
                    del global_source_dict[source.global_path]
        self.source_name_dict = {}
        for source in self.source_list:
            if source.name:
                self.source_name_dict[source.name] = source
