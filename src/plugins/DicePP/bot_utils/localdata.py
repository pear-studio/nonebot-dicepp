import os
import json
import openpyxl


def read_json(path: str) -> dict:
    """
    从path中读取一个json文件并返回对应的字典
    """
    with open(path, "r", encoding='utf-8') as f:
        js = f.read()
        json_dict = json.loads(js)
        return json_dict


def update_json(json_dict: dict, path: str) -> None:
    """
    将jsonFile保存到path路径中
    """
    with open(path, "w", encoding='utf-8') as f:
        json.dump(json_dict, f, ensure_ascii=False)


async def update_json_async(json_dict: dict, path: str) -> None:
    """
    异步地将jsonFile保存到path路径中
    """
    with open(path, "w", encoding='utf-8') as f:
        json.dump(json_dict, f, ensure_ascii=False)


def read_xlsx(path: str) -> openpyxl.Workbook:
    """
    读取xlsx, 为保证workbook正确关闭, 请用类似 with read_xlsx(data_path) as workbook 的方式读取
    """
    return openpyxl.load_workbook(path)


def update_xlsx(workbook: openpyxl.Workbook, path: str) -> None:
    workbook.save(path)


def format_worksheet(sheet, height: float = 30, min_width: float = 20, width_scale: float = 1):
    """
    Args:
        sheet: 要修改的页面, 类型为openpyxl.worksheet.worksheet.Worksheet
        height: 每行高度
        min_width: 每列最小宽度
        width_scale:  列宽=列内最大字符数*width_scale
    """
    from openpyxl.utils import get_column_letter

    for i in range(1, sheet.max_row + 1):
        sheet.row_dimensions[i].height = height
    for i in range(1, sheet.max_column + 1):
        cl = get_column_letter(i)
        width = max((len(cell.value) for cell in sheet[cl] if cell.value)) * width_scale
        width = max(width, min_width)
        sheet.column_dimensions[cl].width = width


def create_parent_dir(path: str):
    """为path递归创建所有不存在的父文件夹"""
    path = os.path.abspath(path)
    assert path and path[0] != "."
    if "." not in path:  # path是文件夹
        dir_name = path
    else:  # path是文件
        dir_name = os.path.dirname(path)

    if not os.path.exists(os.path.dirname(dir_name)):
        create_parent_dir(os.path.dirname(dir_name))

    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
