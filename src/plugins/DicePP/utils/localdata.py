from typing import List, Dict
import os
import json
import openpyxl
from openpyxl.comments import Comment


def read_json(path: str) -> dict:
    """
    从path中读取一个json文件并返回对应的字典
    """
    with open(path, "r", encoding='utf-8') as f:
        js = f.read()
        json_dict = json.loads(js)
        return json_dict


def update_json(json_dict: dict, path: str) -> None:
    """
    将jsonFile保存到path路径中
    """
    with open(path, "w", encoding='utf-8') as f:
        json.dump(json_dict, f, ensure_ascii=False)


async def update_json_async(json_dict: dict, path: str) -> None:
    """
    异步地将jsonFile保存到path路径中
    """
    with open(path, "w", encoding='utf-8') as f:
        json.dump(json_dict, f, ensure_ascii=False)


def read_xlsx(path: str) -> openpyxl.Workbook:
    """
    读取xlsx, 记得之后手动关闭workbook
    """
    wb = openpyxl.load_workbook(path)
    wb_title = path.rsplit("/", maxsplit=1)[-1]
    wb_title = wb_title.rsplit("\\", maxsplit=1)[-1]
    wb_title = wb_title.rsplit(".", maxsplit=1)[0]
    wb.properties.title = wb_title
    wb.properties.identifier = path
    return wb


def update_xlsx(workbook: openpyxl.Workbook, path: str) -> None:
    workbook.save(path)


def get_empty_col_based_workbook(keywords: List[str], keyword_comments: Dict[str, str]) -> openpyxl.Workbook:
    """获得一个模板工作簿"""
    wb = openpyxl.Workbook()
    for name in wb.sheetnames:
        del wb[name]
    ws_temp = wb.create_sheet("template")
    for i, text in enumerate(keywords):
        cell_field = ws_temp.cell(row=1, column=1 + i, value=text)
        cell_field.comment = Comment(keyword_comments[text], "DicePP")
    return wb


def col_based_workbook_to_dict(wb: openpyxl.Workbook, keywords: List[str], error_info: List[str]) -> dict:
    """
    将已经读取的 column based 工作簿转换为dict, 要求第一行是关键字, 后续行是对应内容, 如果sheet中没有任何内容或关键字不完整则不会创建对应的字典
    Args:
        wb: 已经读取的xlsx
        keywords: 关键字列表, 如果为空将会使用表中已有的全部关键字
        error_info: 将检测到的错误加入到error_info中

    Returns:
        result: 字典, 第一级key为sheet名称, 第二级key为关键字, 之后是字符串列表
        通过类似 result[sheet_name][key_name][row_index-2] 来访问对应的数据, 没有数据则返回空字符串
    """
    result = {}
    for sheet_name in wb.sheetnames:
        key_index_dict: Dict[str, int] = {}
        ws = wb[sheet_name]
        # 获取当前关键字列表
        if not keywords:
            keywords_cur = []
            for header_cell in ws[1]:
                keywords_cur.append(header_cell.value)
        else:
            keywords_cur = keywords
        # 获取关键字索引
        for header_cell in ws[1]:
            if header_cell.value in keywords_cur:
                key_index_dict[header_cell.value] = header_cell.column - 1
        # 检测关键字是否完整
        is_valid = True
        for keyword in keywords_cur:
            if keyword not in key_index_dict:
                error_info.append(f"不完整的表格{wb.properties.title}->{sheet_name}, 缺少{keyword}, 未加载该工作表")
                is_valid = False
                break
        if not is_valid:
            continue

        # 生成工作表字典
        result[sheet_name] = {}
        for keyword in keywords_cur:
            result[sheet_name][keyword] = []
        is_valid = False
        # 逐行生成内容
        for row in ws.iter_rows(min_row=2):
            for keyword in keywords_cur:
                content = row[key_index_dict[keyword]].value
                content: str = str(content).strip() if content is not None else ""
                result[sheet_name][keyword].append(content)
                if not is_valid and content:
                    is_valid = True
        if not is_valid:
            error_info.append(f"空工作表{wb.properties.title}->{sheet_name}, 未加载该工作表")
            del result[sheet_name]
    if len(result) == 0:
        error_info.append(f"表格中不含有任何可用工作表{wb.properties.title}")
    return result


def format_worksheet(sheet, height: float = 30, min_width: float = 20, width_scale: float = 1):
    """
    Args:
        sheet: 要修改的页面, 类型为openpyxl.worksheet.worksheet.Worksheet
        height: 每行高度
        min_width: 每列最小宽度
        width_scale:  列宽=列内最大字符数*width_scale
    """
    from openpyxl.utils import get_column_letter

    for i in range(1, sheet.max_row + 1):
        sheet.row_dimensions[i].height = height
    for i in range(1, sheet.max_column + 1):
        cl = get_column_letter(i)
        width = max((len(cell.value) for cell in sheet[cl] if cell.value)) * width_scale
        width = max(width, min_width)
        sheet.column_dimensions[cl].width = width


def create_parent_dir(path: str):
    """为path递归创建所有不存在的父文件夹"""
    path = os.path.abspath(path)
    assert path and path[0] != "."
    if "." not in path:  # path是文件夹
        dir_name = path
    else:  # path是文件
        dir_name = os.path.dirname(path)

    if not os.path.exists(os.path.dirname(dir_name)):
        create_parent_dir(os.path.dirname(dir_name))

    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
