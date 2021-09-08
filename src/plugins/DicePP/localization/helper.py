from typing import Dict
import os

import openpyxl
from openpyxl.comments import Comment

import logger
from bot_utils.localdata import read_xlsx

import localization
from localization.localization_text import LocalizationText

LOCAL_FILE_PATH = "localization.xlsx"


class LocalizationHelper:
    def __init__(self, data_path: str, identifier: str):
        self.data_path = os.path.join(data_path, LOCAL_FILE_PATH)
        self.identifier = identifier
        self.all_local_texts: Dict[str, LocalizationText] = {}

        # 通用的本地化语句
        for key in localization.common_local_text.COMMON_LOCAL_TEXT.keys():
            self.register_loc_text(key,
                                   localization.common_local_text.COMMON_LOCAL_TEXT[key],
                                   localization.common_local_text.COMMON_LOCAL_COMMENT[key])

    def load_localization(self):
        """用文档里的配置覆写之前的配置"""
        if not os.path.exists(self.data_path):
            return
        workbook = read_xlsx(self.data_path)
        if self.identifier in workbook.sheetnames:
            sheet = workbook[self.identifier]
        elif "Default" in workbook.sheetnames:
            sheet = workbook["Default"]
        else:
            return
        for row in sheet.iter_rows():
            key = str(row[0].value)  # 第一个元素为关键字
            if key not in self.all_local_texts:
                continue
            comment: str = self.all_local_texts[key].comment  # 沿用原来的注释, 不用文件里的
            self.all_local_texts[key] = LocalizationText(key, comment=comment)
            for text in [str(cell.value) for cell in row[1:] if cell.value and cell.value.strip()]:
                self.all_local_texts[key].add(text)
        workbook.close()
        logger.dice_log(f"[Localization] [Load] 成功读取本地化文件 {self.data_path}")

    def save_localization(self):
        """按现在的设置多个机器人会读写同一个配置文件, 如果并行可能存在写冲突, 现在应该是单线程异步, 应该没问题"""
        def save_loc_text_to_row(sheet, l_text: LocalizationText, row: int):
            header = sheet.cell(row=row, column=1, value=l_text.key)
            if l_text.comment:
                header.comment = Comment(l_text.comment, "DicePP")
            for ci, text in enumerate(l_text.loc_texts):
                sheet.cell(row=row, column=ci+2, value=text)

        feedback: str
        if os.path.exists(self.data_path):
            workbook = read_xlsx(self.data_path)
            feedback = "成功更新本地化文件"
        else:
            workbook = openpyxl.Workbook()
            for name in workbook.sheetnames:
                del workbook[name]
            feedback = "成功创建本地化文件"
        if self.identifier in workbook.sheetnames:
            cur_sheet = workbook[self.identifier]
        else:
            cur_sheet = workbook.create_sheet(self.identifier)
        for ri, loc_text in enumerate(self.all_local_texts.values()):
            save_loc_text_to_row(cur_sheet, loc_text, ri+1)

        try:
            workbook.save(self.data_path)
        except PermissionError:
            logger.dice_log(f"[Localization] [Save] Save localization {self.data_path} failed as permission denied!")
        workbook.close()

        logger.dice_log(f"[Localization] [Save] {feedback} {self.data_path}")

    def register_loc_text(self, key: str, default_text: str, comment: str = ""):
        """
        将一个本地化语句注册至Helper中
        Args:
            key: 本地化语句的关键字
            default_text: 本地化语句的默认值
            comment: 对本地化语句的注释
        """
        self.all_local_texts[key] = LocalizationText(key, default_text, comment)

    def get_loc_text(self, key: str) -> LocalizationText:
        """
        获取本地化语句
        Args:
            key: 本地化语句的关键字
        """
        return self.all_local_texts[key]

    def format_loc_text(self, key: str, **kwargs) -> str:
        """
        格式化并返回本地化语句
        Args:
            key: 本地化语句的关键字
            **kwargs: 本地化语句需要的参数, 可以传不会用到的参数
        """
        loc_text: LocalizationText = self.get_loc_text(key)
        if kwargs:
            return loc_text.get().format(**kwargs)
        else:
            return loc_text.get()
